from fastapi import FastAPI, UploadFile, File, Body
from fastapi.responses import JSONResponse, FileResponse
from lxml import etree
from html import unescape
from dotenv import load_dotenv
from io import BytesIO
import imaplib
import os
import shutil
import uuid
import zipfile
import sqlite3
import re
import unicodedata

load_dotenv()

app = FastAPI(title="Facturas IA API")

BASE_PATH = os.getenv("BASE_PATH", r"C:\FacturasIA")
ADJUNTOS_PATH = os.path.join(BASE_PATH, "adjuntos")
PROCESADAS_PATH = os.path.join(BASE_PATH, "procesadas")
ERRORES_PATH = os.path.join(BASE_PATH, "errores")
DB_PATH = os.path.join(BASE_PATH, "facturas_ai.db")
AUTO_CLASIFICAR_CONCEPTO_UNICO_PROVEEDOR = (
    os.getenv("AUTO_CLASIFICAR_CONCEPTO_UNICO_PROVEEDOR", "true").strip().lower()
    not in {"0", "false", "no", "off"}
)
SIIGO_AUTO_CREAR_PROVEEDOR = (
    os.getenv("SIIGO_AUTO_CREAR_PROVEEDOR", "true").strip().lower()
    not in {"0", "false", "no", "off"}
)

os.makedirs(ADJUNTOS_PATH, exist_ok=True)
os.makedirs(PROCESADAS_PATH, exist_ok=True)
os.makedirs(ERRORES_PATH, exist_ok=True)


def columna_existe(cursor, tabla: str, columna: str):
    cursor.execute(f"PRAGMA table_info({tabla})")
    return any(row[1] == columna for row in cursor.fetchall())


def inicializar_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nit TEXT UNIQUE,
            nombre TEXT,
            ciudad TEXT,
            regimen TEXT,
            responsabilidad_fiscal TEXT,
            cuenta_gasto_default TEXT,
            cuenta_iva_default TEXT,
            cuenta_retencion_default TEXT,
            activo INTEGER DEFAULT 1,
            fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS facturas_recibidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_factura TEXT,
            fecha_factura TEXT,
            hora_factura TEXT,
            cufe TEXT UNIQUE,
            proveedor_nombre TEXT,
            proveedor_nit TEXT,
            cliente_nombre TEXT,
            cliente_nit TEXT,
            subtotal REAL,
            total_sin_impuestos REAL,
            iva REAL,
            total_con_impuestos REAL,
            total_pagar REAL,
            xml_principal TEXT,
            pdf_principal TEXT,
            archivo_recibido TEXT,
            fecha_procesamiento DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS facturas_lineas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factura_id INTEGER,
            descripcion TEXT,
            cantidad REAL,
            valor_linea REAL,
            concepto_servicio TEXT,
            clasificacion_fuente TEXT,
            clasificacion_confianza REAL,
            FOREIGN KEY (factura_id) REFERENCES facturas_recibidas(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contabilizaciones_historicas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proveedor_nit TEXT,
            proveedor_nombre TEXT,
            descripcion TEXT,
            cuenta_contable TEXT,
            nombre_cuenta TEXT,
            centro_costo TEXT,
            debito REAL DEFAULT 0,
            credito REAL DEFAULT 0,
            cuenta_iva TEXT,
            cuenta_retencion TEXT,
            fuente TEXT,
            fecha_documento TEXT,
            fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS mapeo_erp (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            erp TEXT DEFAULT 'SIIGO',
            proveedor_nit TEXT,
            proveedor_nombre TEXT,
            concepto_servicio TEXT,
            cuenta_contable TEXT,
            nombre_cuenta TEXT,
            item_type_erp TEXT,
            item_code_erp TEXT,
            item_description_erp TEXT,
            document_id_erp TEXT,
            payment_id_erp TEXT,
            tax_id_erp TEXT,
            activo INTEGER DEFAULT 1,
            observacion TEXT,
            fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reglas_concepto_servicio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            erp TEXT DEFAULT 'SIIGO',
            proveedor_nit TEXT,
            concepto_servicio TEXT,
            palabras_clave TEXT,
            cuenta_contable TEXT,
            nombre_cuenta TEXT,
            item_type_erp TEXT,
            item_code_erp TEXT,
            tax_id_erp TEXT,
            prioridad INTEGER DEFAULT 100,
            activo INTEGER DEFAULT 1,
            fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_mapeo_erp_busqueda
        ON mapeo_erp (
            erp,
            proveedor_nit,
            concepto_servicio,
            cuenta_contable,
            activo
        )
    """)

    cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_reglas_concepto_busqueda
        ON reglas_concepto_servicio (
            erp,
            proveedor_nit,
            concepto_servicio,
            activo,
            prioridad
        )
    """)

    if not columna_existe(cursor, "facturas_lineas", "concepto_servicio"):
        cursor.execute("ALTER TABLE facturas_lineas ADD COLUMN concepto_servicio TEXT")

    if not columna_existe(cursor, "facturas_lineas", "clasificacion_fuente"):
        cursor.execute("ALTER TABLE facturas_lineas ADD COLUMN clasificacion_fuente TEXT")

    if not columna_existe(cursor, "facturas_lineas", "clasificacion_confianza"):
        cursor.execute("ALTER TABLE facturas_lineas ADD COLUMN clasificacion_confianza REAL")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS causaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factura_id INTEGER,
            proveedor_nit TEXT,
            proveedor_nombre TEXT,
            numero_factura TEXT,
            estado TEXT,
            confianza REAL,
            total_debito REAL,
            total_credito REAL,
            mensaje TEXT,
            siigo_comprobante_id TEXT,
            fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (factura_id) REFERENCES facturas_recibidas(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS causacion_lineas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            causacion_id INTEGER,
            tipo TEXT,
            cuenta_contable TEXT,
            nombre_cuenta TEXT,
            centro_costo TEXT,
            debito REAL DEFAULT 0,
            credito REAL DEFAULT 0,
            concepto_servicio TEXT,
            descripcion TEXT,
            FOREIGN KEY (causacion_id) REFERENCES causaciones(id)
        )
    """)

    if not columna_existe(cursor, "causacion_lineas", "concepto_servicio"):
        cursor.execute("ALTER TABLE causacion_lineas ADD COLUMN concepto_servicio TEXT")

    conn.commit()
    conn.close()


def buscar_xml_pdf(carpeta: str):
    xml_files = []
    pdf_files = []

    for root_dir, dirs, files in os.walk(carpeta):
        for file_name in files:
            ruta = os.path.join(root_dir, file_name)
            nombre_lower = file_name.lower()

            if nombre_lower.endswith(".xml"):
                xml_files.append(ruta)

            if nombre_lower.endswith(".pdf"):
                pdf_files.append(ruta)

    return xml_files, pdf_files


def normalizar_clave_excel(valor):
    if valor is None:
        return ""

    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def valor_texto_excel(valor):
    if valor is None:
        return None

    if isinstance(valor, str):
        texto = valor.strip()
        return texto or None

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip() or None


def valor_numero_excel(valor, default=0):
    if valor is None or valor == "":
        return default

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return default

    texto = re.sub(r"[^0-9,.\-]", "", texto)

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return default


def valor_fecha_excel(valor):
    if valor is None:
        return None

    if hasattr(valor, "date"):
        return valor.date().isoformat()

    return valor_texto_excel(valor)


def valor_activo_excel(valor, default=1):
    if valor is None or valor == "":
        return default

    texto = normalizar_clave_excel(valor)

    if texto in {"1", "si", "s", "true", "activo", "activa", "yes", "y"}:
        return 1

    if texto in {"0", "no", "false", "inactivo", "inactiva", "n"}:
        return 0

    return default


def obtener_valor_fila(fila, *alias):
    for nombre in alias:
        clave = normalizar_clave_excel(nombre)
        if clave in fila:
            valor = fila[clave]
            if valor is not None and valor != "":
                return valor

    return None


def obtener_hoja_excel(workbook, *alias):
    hojas_por_nombre = {
        normalizar_clave_excel(nombre): nombre
        for nombre in workbook.sheetnames
    }

    for nombre in alias:
        clave = normalizar_clave_excel(nombre)
        if clave in hojas_por_nombre:
            return workbook[hojas_por_nombre[clave]]

    return None


def leer_filas_excel(worksheet):
    filas = worksheet.iter_rows(values_only=True)

    for fila_encabezado in filas:
        encabezados = [normalizar_clave_excel(valor) for valor in fila_encabezado]

        if any(encabezados):
            break
    else:
        return []

    registros = []

    for fila in filas:
        registro = {}

        for indice, encabezado in enumerate(encabezados):
            if not encabezado:
                continue

            valor = fila[indice] if indice < len(fila) else None
            registro[encabezado] = valor

        if any(valor not in (None, "") for valor in registro.values()):
            registros.append(registro)

    return registros


def construir_historico_desde_excel(fila, nombre_archivo):
    proveedor_nit = valor_texto_excel(obtener_valor_fila(
        fila,
        "proveedor_nit",
        "nit_proveedor",
        "nit",
        "tercero_nit",
        "identificacion",
        "identificacion_tercero"
    ))
    cuenta_contable = valor_texto_excel(obtener_valor_fila(
        fila,
        "cuenta_contable",
        "codigo_cuenta",
        "cuenta_codigo",
        "cuenta"
    ))

    if not proveedor_nit or not cuenta_contable:
        return None

    return {
        "proveedor_nit": proveedor_nit,
        "proveedor_nombre": valor_texto_excel(obtener_valor_fila(
            fila,
            "proveedor_nombre",
            "nombre_proveedor",
            "proveedor",
            "tercero",
            "nombre_tercero",
            "nombre"
        )),
        "descripcion": valor_texto_excel(obtener_valor_fila(
            fila,
            "descripcion",
            "detalle",
            "concepto",
            "observacion",
            "producto_servicio"
        )),
        "cuenta_contable": cuenta_contable,
        "nombre_cuenta": valor_texto_excel(obtener_valor_fila(
            fila,
            "nombre_cuenta",
            "cuenta_nombre",
            "descripcion_cuenta"
        )),
        "centro_costo": valor_texto_excel(obtener_valor_fila(
            fila,
            "centro_costo",
            "centro_de_costo",
            "codigo_centro_costo",
            "cc"
        )),
        "debito": valor_numero_excel(obtener_valor_fila(
            fila,
            "debito",
            "debitos",
            "valor_debito"
        )),
        "credito": valor_numero_excel(obtener_valor_fila(
            fila,
            "credito",
            "creditos",
            "valor_credito"
        )),
        "cuenta_iva": valor_texto_excel(obtener_valor_fila(
            fila,
            "cuenta_iva",
            "iva_cuenta"
        )),
        "cuenta_retencion": valor_texto_excel(obtener_valor_fila(
            fila,
            "cuenta_retencion",
            "retencion_cuenta",
            "cuenta_retefuente"
        )),
        "fuente": valor_texto_excel(obtener_valor_fila(
            fila,
            "fuente",
            "origen"
        )) or f"Excel: {nombre_archivo}",
        "fecha_documento": valor_fecha_excel(obtener_valor_fila(
            fila,
            "fecha_documento",
            "fecha",
            "fecha_comprobante",
            "fecha_factura"
        ))
    }


def construir_mapeo_desde_excel(fila):
    proveedor_nit = valor_texto_excel(obtener_valor_fila(
        fila,
        "proveedor_nit",
        "nit_proveedor",
        "nit",
        "tercero_nit"
    ))
    concepto_servicio = valor_texto_excel(obtener_valor_fila(
        fila,
        "concepto_servicio",
        "concepto",
        "servicio",
        "producto_servicio",
        "categoria"
    ))
    cuenta_contable = valor_texto_excel(obtener_valor_fila(
        fila,
        "cuenta_contable",
        "codigo_cuenta",
        "cuenta_codigo",
        "cuenta"
    ))
    item_code_erp = valor_texto_excel(obtener_valor_fila(
        fila,
        "item_code_erp",
        "codigo_item_erp",
        "item_code",
        "code",
        "codigo_producto",
        "codigo"
    ))

    if not proveedor_nit or not (concepto_servicio or cuenta_contable or item_code_erp):
        return None

    nombre_cuenta = valor_texto_excel(obtener_valor_fila(
        fila,
        "nombre_cuenta",
        "cuenta_nombre",
        "descripcion_cuenta"
    ))
    item_description_erp = valor_texto_excel(obtener_valor_fila(
        fila,
        "item_description_erp",
        "descripcion_item_erp",
        "item_description",
        "description",
        "descripcion_producto"
    ))

    concepto_servicio = concepto_servicio or inferir_concepto_servicio(
        item_description_erp,
        nombre_cuenta
    )

    return {
        "erp": valor_texto_excel(obtener_valor_fila(fila, "erp")) or "SIIGO",
        "proveedor_nit": proveedor_nit,
        "proveedor_nombre": valor_texto_excel(obtener_valor_fila(
            fila,
            "proveedor_nombre",
            "nombre_proveedor",
            "proveedor",
            "tercero",
            "nombre_tercero",
            "nombre"
        )),
        "concepto_servicio": concepto_servicio,
        "cuenta_contable": cuenta_contable,
        "nombre_cuenta": nombre_cuenta,
        "item_type_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "item_type_erp",
            "tipo_item_erp",
            "tipo_erp",
            "item_type",
            "type"
        )),
        "item_code_erp": item_code_erp,
        "item_description_erp": item_description_erp,
        "document_id_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "document_id_erp",
            "documento_erp",
            "document_id"
        )),
        "payment_id_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "payment_id_erp",
            "medio_pago_erp",
            "payment_id"
        )),
        "tax_id_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "tax_id_erp",
            "impuesto_erp",
            "tax_id"
        )),
        "activo": valor_activo_excel(obtener_valor_fila(fila, "activo")),
        "observacion": valor_texto_excel(obtener_valor_fila(
            fila,
            "observacion",
            "observaciones",
            "nota"
        ))
    }


def construir_regla_desde_excel(fila):
    proveedor_nit = valor_texto_excel(obtener_valor_fila(
        fila,
        "proveedor_nit",
        "nit_proveedor",
        "nit",
        "tercero_nit"
    ))
    concepto_servicio = valor_texto_excel(obtener_valor_fila(
        fila,
        "concepto_servicio",
        "concepto",
        "servicio",
        "producto_servicio",
        "categoria"
    ))
    palabras_clave = valor_texto_excel(obtener_valor_fila(
        fila,
        "palabras_clave",
        "keywords",
        "palabras",
        "regla"
    ))

    if not proveedor_nit or not concepto_servicio or not palabras_clave:
        return None

    return {
        "erp": valor_texto_excel(obtener_valor_fila(fila, "erp")) or "SIIGO",
        "proveedor_nit": proveedor_nit,
        "concepto_servicio": concepto_servicio,
        "palabras_clave": palabras_clave,
        "cuenta_contable": valor_texto_excel(obtener_valor_fila(
            fila,
            "cuenta_contable",
            "codigo_cuenta",
            "cuenta_codigo",
            "cuenta"
        )),
        "nombre_cuenta": valor_texto_excel(obtener_valor_fila(
            fila,
            "nombre_cuenta",
            "cuenta_nombre",
            "descripcion_cuenta"
        )),
        "item_type_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "item_type_erp",
            "tipo_item_erp",
            "tipo_erp",
            "item_type",
            "type"
        )),
        "item_code_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "item_code_erp",
            "codigo_item_erp",
            "item_code",
            "code",
            "codigo_producto",
            "codigo"
        )),
        "tax_id_erp": valor_texto_excel(obtener_valor_fila(
            fila,
            "tax_id_erp",
            "impuesto_erp",
            "tax_id"
        )),
        "prioridad": int(valor_numero_excel(obtener_valor_fila(fila, "prioridad"), 100)),
        "activo": valor_activo_excel(obtener_valor_fila(fila, "activo"))
    }


def normalizar_texto_busqueda(valor):
    return normalizar_clave_excel(valor).replace("_", " ")


def dividir_palabras_clave(valor):
    if not valor:
        return []

    partes = re.split(r"[,;|\n]+", str(valor))
    palabras = []

    for parte in partes:
        texto = normalizar_texto_busqueda(parte)

        if texto and texto not in palabras:
            palabras.append(texto)

    return palabras


def generar_palabras_desde_concepto(concepto):
    texto = normalizar_texto_busqueda(concepto)

    if not texto:
        return []

    palabras = [texto]

    for token in texto.split():
        if len(token) >= 3 and token not in palabras:
            palabras.append(token)

    return palabras


PALABRAS_APRENDIZAJE_IGNORADAS = {
    "para",
    "por",
    "con",
    "del",
    "las",
    "los",
    "una",
    "uno",
    "sus",
    "sas",
    "s.a.s",
    "sa",
    "s.a",
    "ltda",
    "colombia",
    "factura",
    "servicio",
    "servicios",
    "producto",
    "productos",
    "unidad",
    "cantidad"
}


def generar_palabras_clave_descripcion_aprendida(descripcion):
    texto = normalizar_texto_busqueda(descripcion)

    if not texto:
        return []

    texto = re.sub(r"[^a-z0-9]+", " ", texto).strip()

    if not texto:
        return []

    tokens_originales = texto.split()
    tokens_filtrados = []

    for token in tokens_originales:
        if token in PALABRAS_APRENDIZAJE_IGNORADAS:
            continue

        if token.isdigit() and len(token) > 4:
            continue

        if len(token) > 18 and re.search(r"[a-z]", token) and re.search(r"\d", token):
            continue

        if len(token) < 4 and not token.isdigit():
            continue

        tokens_filtrados.append(token)

    palabras = []

    def agregar(valor):
        valor = " ".join(str(valor or "").split()).strip()

        if valor and valor not in palabras:
            palabras.append(valor)

    for indice, token in enumerate(tokens_originales):
        if token in {"microsoft", "office", "google", "adobe"} and indice + 1 < len(tokens_originales):
            siguiente = tokens_originales[indice + 1]
            if siguiente.isdigit() or len(siguiente) >= 3:
                agregar(f"{token} {siguiente}")

    for tamano in (3, 2):
        for indice in range(0, max(0, len(tokens_filtrados) - tamano + 1)):
            agregar(" ".join(tokens_filtrados[indice:indice + tamano]))

    for token in tokens_filtrados:
        agregar(token)

    return palabras[:12]


CONCEPTOS_SERVICIO_KEYWORDS = [
    ("CERTIFICADO_SSL", ["certificado ssl", "ssl", "tls", "https"]),
    ("HOSTING", ["hosting", "hospedaje", "web hosting", "alojamiento web"]),
    ("DOMINIO", ["dominio", "domain", "renovacion dominio"]),
    ("CORREO", ["correo", "email", "e mail", "mail", "google workspace"]),
    ("LICENCIA", [
        "licencia",
        "license",
        "suscripcion",
        "subscription",
        "microsoft 365",
        "office 365",
        "business standard",
        "m365",
        "software assurance"
    ]),
    ("SERVIDOR", ["servidor", "server", "vps", "cloud", "nube"]),
    ("DESARROLLO_SOFTWARE", ["desarrollo software", "software", "programacion", "aplicacion"]),
    ("SERVICIOS_TI", ["servicios ti", "consultoria ti", "tecnologia", "sistemas"]),
    ("SERVICIOS_PUBLICOS", [
        "servicios publicos",
        "energia",
        "energia domiciliario",
        "acueducto",
        "agua",
        "aseo",
        "alcantarillado",
        "alumbrado",
        "alumbrado publico",
        "tasa aseo",
        "tasa"
    ]),
    ("TELECOMUNICACIONES", [
        "internet",
        "fibra",
        "telefonia",
        "telecomunicaciones",
        "television",
        "televison",
        "tv",
        "cable",
        "television por cable"
    ]),
    ("MANTENIMIENTO_EQUIPO", ["mantenimiento", "equipo computacion", "computador"]),
    ("PAPELERIA", ["papeleria", "utiles", "fotocopias"]),
    ("COMBUSTIBLE", ["combustible", "lubricante", "gasolina", "diesel"]),
    ("REGISTRO_MERCANTIL", ["registro mercantil", "camara de comercio"]),
    ("EDUCATIVO", ["educativo", "capacitacion", "curso", "universidad"]),
    ("ASESORIA", ["asesoria", "consultoria"]),
]


def inferir_concepto_servicio(*textos):
    texto = " ".join(normalizar_texto_busqueda(t) for t in textos if t)

    if not texto:
        return None

    for concepto, palabras in CONCEPTOS_SERVICIO_KEYWORDS:
        if any(normalizar_texto_busqueda(palabra) in texto for palabra in palabras):
            return concepto

    return None


def conceptos_compatibles(concepto_solicitado, concepto_mapeo):
    if not concepto_solicitado:
        return True

    if not concepto_mapeo:
        return True

    if concepto_solicitado == concepto_mapeo:
        return True

    grupo_ti = {
        "CERTIFICADO_SSL",
        "HOSTING",
        "DOMINIO",
        "CORREO",
        "LICENCIA",
        "SERVIDOR",
        "DESARROLLO_SOFTWARE",
        "SERVICIOS_TI"
    }

    if concepto_solicitado in grupo_ti and concepto_mapeo in grupo_ti:
        return True

    grupo_servicios_publicos = {
        "SERVICIOS_PUBLICOS",
        "ENERGIA",
        "ACUEDUCTO",
        "ASEO",
        "ALCANTARILLADO",
        "ALUMBRADO_PUBLICO"
    }

    if concepto_solicitado in grupo_servicios_publicos and concepto_mapeo in grupo_servicios_publicos:
        return True

    return False


def calcular_match_palabras(descripcion_normalizada, palabras):
    coincidencias = []

    for palabra in palabras:
        palabra_normalizada = normalizar_texto_busqueda(palabra)

        if not palabra_normalizada:
            continue

        if palabra_normalizada in descripcion_normalizada:
            coincidencias.append(palabra_normalizada)

    return coincidencias


def cargar_reglas_clasificacion(cursor, proveedor_nit):
    cursor.execute("""
        SELECT
            id,
            erp,
            proveedor_nit,
            concepto_servicio,
            palabras_clave,
            cuenta_contable,
            nombre_cuenta,
            item_type_erp,
            item_code_erp,
            tax_id_erp,
            prioridad
        FROM reglas_concepto_servicio
        WHERE activo = 1
          AND (
                proveedor_nit = ?
                OR proveedor_nit IS NULL
                OR proveedor_nit = ''
          )
        ORDER BY
            CASE
                WHEN proveedor_nit = ? THEN 0
                ELSE 1
            END,
            prioridad ASC,
            id ASC
    """, (proveedor_nit, proveedor_nit))

    return [dict(row) for row in cursor.fetchall()]


def cargar_mapeos_clasificacion(cursor, proveedor_nit):
    cursor.execute("""
        SELECT
            id,
            erp,
            proveedor_nit,
            proveedor_nombre,
            concepto_servicio,
            cuenta_contable,
            nombre_cuenta,
            item_type_erp,
            item_code_erp,
            item_description_erp,
            tax_id_erp,
            observacion
        FROM mapeo_erp
        WHERE activo = 1
          AND proveedor_nit = ?
        ORDER BY id ASC
    """, (proveedor_nit,))

    mapeos = []

    for row in cursor.fetchall():
        mapeo = dict(row)

        if not mapeo.get("concepto_servicio"):
            mapeo["concepto_servicio"] = inferir_concepto_servicio(
                mapeo.get("item_description_erp"),
                mapeo.get("nombre_cuenta"),
                mapeo.get("observacion")
            )

        mapeos.append(mapeo)

    return mapeos


def buscar_mapeo_erp(cursor, proveedor_nit, concepto_servicio=None, cuenta_contable=None):
    cursor.execute("""
        SELECT *
        FROM mapeo_erp
        WHERE activo = 1
          AND proveedor_nit = ?
        ORDER BY
            CASE
                WHEN concepto_servicio = ? THEN 0
                WHEN concepto_servicio IS NULL OR TRIM(concepto_servicio) = '' THEN 1
                ELSE 2
            END,
            CASE
                WHEN cuenta_contable = ? THEN 0
                WHEN ? IS NULL THEN 1
                ELSE 2
            END,
            id ASC
    """, (proveedor_nit, concepto_servicio, cuenta_contable, cuenta_contable))

    candidatos = []

    for row in cursor.fetchall():
        mapeo = dict(row)
        concepto_mapeo = mapeo.get("concepto_servicio") or inferir_concepto_servicio(
            mapeo.get("item_description_erp"),
            mapeo.get("nombre_cuenta"),
            mapeo.get("observacion")
        )

        if not conceptos_compatibles(concepto_servicio, concepto_mapeo):
            continue

        if cuenta_contable and mapeo.get("cuenta_contable") and mapeo.get("cuenta_contable") != cuenta_contable:
            continue

        mapeo["concepto_servicio_resuelto"] = concepto_mapeo
        candidatos.append(mapeo)

    if not candidatos:
        return None

    if concepto_servicio:
        exactos = [
            m for m in candidatos
            if m.get("concepto_servicio_resuelto") == concepto_servicio
        ]
        if exactos:
            return exactos[0]

    if len(candidatos) == 1:
        return candidatos[0]

    genericos = [
        m for m in candidatos
        if not m.get("concepto_servicio") and m.get("item_type_erp") and m.get("item_code_erp")
    ]

    return genericos[0] if len(genericos) == 1 else None


def clasificar_linea_por_reglas(descripcion, reglas):
    descripcion_normalizada = normalizar_texto_busqueda(descripcion)

    for regla in reglas:
        palabras = dividir_palabras_clave(regla.get("palabras_clave"))
        coincidencias = calcular_match_palabras(descripcion_normalizada, palabras)

        if coincidencias:
            return {
                "concepto_servicio": regla.get("concepto_servicio"),
                "clasificacion_fuente": "reglas_concepto_servicio",
                "clasificacion_confianza": 100,
                "regla_id": regla.get("id"),
                "mapeo_erp_id": None,
                "coincidencias": coincidencias
            }

    return None


def clasificar_linea_por_patrones_base(descripcion):
    concepto = inferir_concepto_servicio(descripcion)

    if not concepto:
        return None

    return {
        "concepto_servicio": concepto,
        "clasificacion_fuente": "patrones_base",
        "clasificacion_confianza": 75,
        "regla_id": None,
        "mapeo_erp_id": None,
        "coincidencias": [concepto]
    }


def clasificar_linea_por_mapeo(descripcion, mapeos):
    descripcion_normalizada = normalizar_texto_busqueda(descripcion)
    candidatos = []

    for mapeo in mapeos:
        palabras = []
        palabras.extend(generar_palabras_desde_concepto(mapeo.get("concepto_servicio")))
        palabras.extend(dividir_palabras_clave(mapeo.get("item_description_erp")))
        palabras.extend(dividir_palabras_clave(mapeo.get("observacion")))

        nombre_cuenta = normalizar_texto_busqueda(mapeo.get("nombre_cuenta"))
        if nombre_cuenta:
            palabras.append(nombre_cuenta)

        palabras_unicas = []

        for palabra in palabras:
            if palabra and palabra not in palabras_unicas:
                palabras_unicas.append(palabra)

        coincidencias = calcular_match_palabras(descripcion_normalizada, palabras_unicas)

        if not coincidencias:
            continue

        tiene_concepto_exacto = (
            normalizar_texto_busqueda(mapeo.get("concepto_servicio"))
            in coincidencias
        )
        confianza = 85 if tiene_concepto_exacto else 70

        candidatos.append({
            "concepto_servicio": mapeo.get("concepto_servicio"),
            "clasificacion_fuente": "mapeo_erp",
            "clasificacion_confianza": confianza,
            "regla_id": None,
            "mapeo_erp_id": mapeo.get("id"),
            "coincidencias": coincidencias
        })

    if not candidatos:
        return None

    candidatos.sort(
        key=lambda c: (
            c["clasificacion_confianza"],
            len(c["coincidencias"])
        ),
        reverse=True
    )

    return candidatos[0]


def clasificar_linea_por_concepto_unico_proveedor(descripcion, mapeos):
    if not AUTO_CLASIFICAR_CONCEPTO_UNICO_PROVEEDOR:
        return None

    if not descripcion:
        return None

    conceptos_no_heredables = {"IVA", "CXP"}
    candidatos_por_concepto = {}

    for mapeo in mapeos:
        concepto = mapeo.get("concepto_servicio")

        if not concepto or concepto in conceptos_no_heredables:
            continue

        if not mapeo.get("item_type_erp") or not mapeo.get("item_code_erp"):
            continue

        candidatos_por_concepto.setdefault(concepto, []).append(mapeo)

    if len(candidatos_por_concepto) != 1:
        return None

    concepto = next(iter(candidatos_por_concepto.keys()))
    mapeo = candidatos_por_concepto[concepto][0]

    return {
        "concepto_servicio": concepto,
        "clasificacion_fuente": "mapeo_erp_concepto_unico_proveedor",
        "clasificacion_confianza": 60,
        "regla_id": None,
        "mapeo_erp_id": mapeo.get("id"),
        "coincidencias": ["concepto_unico_proveedor"]
    }


def clasificar_lineas_factura(factura_id: int, guardar: bool = True):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM facturas_recibidas
        WHERE id = ?
    """, (factura_id,))

    factura_row = cursor.fetchone()

    if not factura_row:
        conn.close()
        return {
            "ok": False,
            "status_code": 404,
            "error": f"No existe la factura {factura_id}."
        }

    factura = dict(factura_row)
    proveedor_nit = factura.get("proveedor_nit")

    cursor.execute("""
        SELECT
            id,
            factura_id,
            descripcion,
            cantidad,
            valor_linea,
            concepto_servicio,
            clasificacion_fuente,
            clasificacion_confianza
        FROM facturas_lineas
        WHERE factura_id = ?
        ORDER BY id
    """, (factura_id,))

    lineas = [dict(row) for row in cursor.fetchall()]

    reglas = cargar_reglas_clasificacion(cursor, proveedor_nit)
    mapeos = cargar_mapeos_clasificacion(cursor, proveedor_nit)

    clasificaciones = []
    pendientes = []

    for linea in lineas:
        clasificacion = clasificar_linea_por_reglas(linea.get("descripcion"), reglas)

        if not clasificacion:
            clasificacion = clasificar_linea_por_patrones_base(linea.get("descripcion"))

        if not clasificacion:
            clasificacion = clasificar_linea_por_mapeo(linea.get("descripcion"), mapeos)

        if not clasificacion:
            clasificacion = clasificar_linea_por_concepto_unico_proveedor(
                linea.get("descripcion"),
                mapeos
            )

        if not clasificacion:
            clasificacion = {
                "concepto_servicio": None,
                "clasificacion_fuente": None,
                "clasificacion_confianza": 0,
                "regla_id": None,
                "mapeo_erp_id": None,
                "coincidencias": []
            }
            pendientes.append(linea.get("id"))

        resultado_linea = {
            "linea_id": linea.get("id"),
            "descripcion": linea.get("descripcion"),
            "cantidad": linea.get("cantidad"),
            "valor_linea": linea.get("valor_linea"),
            **clasificacion
        }
        clasificaciones.append(resultado_linea)

        if guardar:
            cursor.execute("""
                UPDATE facturas_lineas
                SET
                    concepto_servicio = ?,
                    clasificacion_fuente = ?,
                    clasificacion_confianza = ?
                WHERE id = ?
            """, (
                clasificacion.get("concepto_servicio"),
                clasificacion.get("clasificacion_fuente"),
                clasificacion.get("clasificacion_confianza"),
                linea.get("id")
            ))

    if guardar:
        conn.commit()

    conn.close()

    return {
        "ok": True,
        "factura_id": factura_id,
        "proveedor_nit": proveedor_nit,
        "proveedor_nombre": factura.get("proveedor_nombre"),
        "numero_factura": factura.get("numero_factura"),
        "guardar": guardar,
        "total_lineas": len(lineas),
        "lineas_clasificadas": len(lineas) - len(pendientes),
        "lineas_pendientes": len(pendientes),
        "requiere_revision_concepto": len(pendientes) > 0,
        "estado": "REQUIERE_REVISION_CONCEPTO" if pendientes else "CONCEPTOS_CLASIFICADOS",
        "total_reglas_disponibles": len(reglas),
        "total_mapeos_disponibles": len(mapeos),
        "clasificaciones": clasificaciones
    }


def safe_extract_zip(zip_ref: zipfile.ZipFile, destino: str):
    destino_abs = os.path.abspath(destino)

    for member in zip_ref.infolist():
        member_path = os.path.abspath(os.path.join(destino, member.filename))

        if not member_path.startswith(destino_abs):
            raise ValueError("ZIP contiene rutas no permitidas.")

    zip_ref.extractall(destino)


def cargar_xml_factura(xml_path: str):
    parser = etree.XMLParser(recover=True, huge_tree=True)
    tree = etree.parse(xml_path, parser)
    root = tree.getroot()

    tag_root = etree.QName(root).localname

    if tag_root == "Invoice":
        return root

    ns = {
        "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
    }

    descripcion = root.findtext(
        ".//cac:Attachment//cac:ExternalReference//cbc:Description",
        namespaces=ns
    )

    if descripcion and "<Invoice" in descripcion:
        return etree.fromstring(descripcion.encode("utf-8"), parser=parser)

    raise ValueError(
        f"No se encontrÃ³ una factura tipo Invoice en el XML. Tipo raÃ­z encontrado: {tag_root}"
    )


def texto(root, xpath: str, ns: dict):
    valor = root.findtext(xpath, namespaces=ns)
    return valor.strip() if valor else None


def numero_decimal(valor):
    try:
        return float(valor) if valor not in [None, ""] else 0
    except Exception:
        return 0


def normalizar_nit(valor):
    return re.sub(r"\D+", "", str(valor or ""))


def obtener_nits_cliente_permitidos():
    valor = os.getenv("NITS_CLIENTE_PERMITIDOS", "")
    partes = re.split(r"[,;\n]+", valor)
    return [normalizar_nit(parte) for parte in partes if normalizar_nit(parte)]


def nit_coincide_permitido(nit_cliente, nit_permitido):
    cliente = normalizar_nit(nit_cliente)
    permitido = normalizar_nit(nit_permitido)

    if not cliente or not permitido:
        return False

    if cliente == permitido:
        return True

    if len(cliente) == len(permitido) + 1 and cliente.startswith(permitido):
        return True

    if len(permitido) == len(cliente) + 1 and permitido.startswith(cliente):
        return True

    return False


def validar_nit_cliente_permitido(datos_factura):
    nits_permitidos = obtener_nits_cliente_permitidos()

    if not nits_permitidos:
        return {
            "permitido": True,
            "validacion_activa": False,
            "nits_permitidos": []
        }

    cliente_nit = datos_factura.get("cliente_nit")
    cliente_nit_normalizado = normalizar_nit(cliente_nit)
    permitido = any(
        nit_coincide_permitido(cliente_nit_normalizado, nit_permitido)
        for nit_permitido in nits_permitidos
    )

    return {
        "permitido": permitido,
        "validacion_activa": True,
        "cliente_nit": cliente_nit,
        "cliente_nit_normalizado": cliente_nit_normalizado,
        "nits_permitidos": nits_permitidos
    }


def texto_atributo(element, xpath, atributo, ns):
    nodo = element.find(xpath, namespaces=ns)

    if nodo is None:
        return None

    valor = nodo.get(atributo)

    if valor is None:
        return None

    valor = str(valor).strip()
    return valor or None


def extraer_datos_xml(xml_path: str):
    root = cargar_xml_factura(xml_path)

    ns = {
        "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
    }

    numero_factura = texto(root, "./cbc:ID", ns)
    fecha_factura = texto(root, "./cbc:IssueDate", ns)
    hora_factura = texto(root, "./cbc:IssueTime", ns)
    cufe = texto(root, "./cbc:UUID", ns)

    proveedor_nombre = texto(
        root,
        ".//cac:AccountingSupplierParty//cac:Party//cac:PartyLegalEntity//cbc:RegistrationName",
        ns
    )

    if not proveedor_nombre:
        proveedor_nombre = texto(
            root,
            ".//cac:AccountingSupplierParty//cac:Party//cac:PartyName//cbc:Name",
            ns
        )

    proveedor_nit = texto(
        root,
        ".//cac:AccountingSupplierParty//cac:Party//cac:PartyTaxScheme//cbc:CompanyID",
        ns
    )
    proveedor_dv = texto_atributo(
        root,
        ".//cac:AccountingSupplierParty//cac:Party//cac:PartyTaxScheme//cbc:CompanyID",
        "schemeID",
        ns
    )

    cliente_nombre = texto(
        root,
        ".//cac:AccountingCustomerParty//cac:Party//cac:PartyLegalEntity//cbc:RegistrationName",
        ns
    )

    cliente_nit = texto(
        root,
        ".//cac:AccountingCustomerParty//cac:Party//cac:PartyTaxScheme//cbc:CompanyID",
        ns
    )

    subtotal = numero_decimal(
        texto(root, ".//cac:LegalMonetaryTotal//cbc:LineExtensionAmount", ns)
    )

    total_sin_impuestos = numero_decimal(
        texto(root, ".//cac:LegalMonetaryTotal//cbc:TaxExclusiveAmount", ns)
    )

    total_con_impuestos = numero_decimal(
        texto(root, ".//cac:LegalMonetaryTotal//cbc:TaxInclusiveAmount", ns)
    )

    total_pagar = numero_decimal(
        texto(root, ".//cac:LegalMonetaryTotal//cbc:PayableAmount", ns)
    )

    iva = 0
    tax_totals = root.findall("./cac:TaxTotal", namespaces=ns)

    for tax_total in tax_totals:
        tax_id = tax_total.findtext(".//cac:TaxScheme/cbc:ID", namespaces=ns)
        tax_amount = tax_total.findtext("./cbc:TaxAmount", namespaces=ns)

        if tax_id == "01":
            iva += numero_decimal(tax_amount)

    lineas = []
    invoice_lines = root.findall(".//cac:InvoiceLine", namespaces=ns)

    for line in invoice_lines:
        descripcion = texto(line, ".//cac:Item/cbc:Description", ns)
        cantidad = numero_decimal(texto(line, "./cbc:InvoicedQuantity", ns))
        valor_linea = numero_decimal(texto(line, "./cbc:LineExtensionAmount", ns))

        lineas.append({
            "descripcion": descripcion,
            "cantidad": cantidad,
            "valor_linea": valor_linea
        })

    return {
        "numero_factura": numero_factura,
        "fecha_factura": fecha_factura,
        "hora_factura": hora_factura,
        "cufe": cufe,
        "proveedor_nombre": proveedor_nombre,
        "proveedor_nit": proveedor_nit,
        "proveedor_dv": proveedor_dv,
        "cliente_nombre": cliente_nombre,
        "cliente_nit": cliente_nit,
        "subtotal": subtotal,
        "total_sin_impuestos": total_sin_impuestos,
        "iva": iva,
        "total_con_impuestos": total_con_impuestos,
        "total_pagar": total_pagar,
        "lineas": lineas
    }


def registrar_proveedor(datos_factura):
    proveedor_nit = datos_factura.get("proveedor_nit")
    proveedor_nombre = datos_factura.get("proveedor_nombre")

    if not proveedor_nit:
        return {
            "proveedor_id": None,
            "proveedor_nuevo": False
        }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM proveedores
        WHERE nit = ?
    """, (proveedor_nit,))

    row = cursor.fetchone()

    if row:
        conn.close()
        return {
            "proveedor_id": row[0],
            "proveedor_nuevo": False
        }

    cursor.execute("""
        INSERT INTO proveedores (
            nit,
            nombre
        )
        VALUES (?, ?)
    """, (
        proveedor_nit,
        proveedor_nombre
    ))

    proveedor_id = cursor.lastrowid

    conn.commit()
    conn.close()

    return {
        "proveedor_id": proveedor_id,
        "proveedor_nuevo": True
    }


def guardar_factura_db(datos_factura, xml_principal, pdf_principal, archivo_recibido):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO facturas_recibidas (
                numero_factura,
                fecha_factura,
                hora_factura,
                cufe,
                proveedor_nombre,
                proveedor_nit,
                cliente_nombre,
                cliente_nit,
                subtotal,
                total_sin_impuestos,
                iva,
                total_con_impuestos,
                total_pagar,
                xml_principal,
                pdf_principal,
                archivo_recibido
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datos_factura.get("numero_factura"),
            datos_factura.get("fecha_factura"),
            datos_factura.get("hora_factura"),
            datos_factura.get("cufe"),
            datos_factura.get("proveedor_nombre"),
            datos_factura.get("proveedor_nit"),
            datos_factura.get("cliente_nombre"),
            datos_factura.get("cliente_nit"),
            datos_factura.get("subtotal"),
            datos_factura.get("total_sin_impuestos"),
            datos_factura.get("iva"),
            datos_factura.get("total_con_impuestos"),
            datos_factura.get("total_pagar"),
            xml_principal,
            pdf_principal,
            archivo_recibido
        ))

        factura_id = cursor.lastrowid

        for linea in datos_factura.get("lineas", []):
            cursor.execute("""
                INSERT INTO facturas_lineas (
                    factura_id,
                    descripcion,
                    cantidad,
                    valor_linea
                )
                VALUES (?, ?, ?, ?)
            """, (
                factura_id,
                linea.get("descripcion"),
                linea.get("cantidad"),
                linea.get("valor_linea")
            ))

        conn.commit()

        return {
            "factura_id": factura_id,
            "duplicada": False
        }

    except sqlite3.IntegrityError:
        cufe = datos_factura.get("cufe")

        cursor.execute("""
            SELECT id
            FROM facturas_recibidas
            WHERE cufe = ?
        """, (cufe,))

        row = cursor.fetchone()
        factura_id = row[0] if row else None

        return {
            "factura_id": factura_id,
            "duplicada": True
        }

    finally:
        conn.close()


def construir_propuesta_causacion(factura_id: int):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM facturas_recibidas
        WHERE id = ?
    """, (factura_id,))

    factura_row = cursor.fetchone()

    if not factura_row:
        conn.close()
        return {
            "ok": False,
            "status_code": 404,
            "error": "Factura no encontrada"
        }

    factura = dict(factura_row)
    proveedor_nit = factura.get("proveedor_nit")
    conn.close()

    total_pagar = round(float(factura.get("total_pagar") or 0), 2)
    iva = round(float(factura.get("iva") or 0), 2)
    subtotal = round(float(factura.get("subtotal") or 0), 2)
    total_sin_impuestos = round(float(factura.get("total_sin_impuestos") or 0), 2)

    if total_pagar > 0:
        valor_gasto = round(total_pagar - iva, 2)
    elif subtotal > 0:
        valor_gasto = subtotal
    elif total_sin_impuestos > 0:
        valor_gasto = total_sin_impuestos
    else:
        valor_gasto = 0

    clasificacion = clasificar_lineas_factura(factura_id, guardar=True)

    if not clasificacion.get("ok"):
        return clasificacion

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM contabilizaciones_historicas
        WHERE proveedor_nit = ?
        ORDER BY fecha_documento DESC, id DESC
        LIMIT 20
    """, (proveedor_nit,))

    historico = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
        SELECT
            id,
            descripcion,
            cantidad,
            valor_linea,
            concepto_servicio,
            clasificacion_fuente,
            clasificacion_confianza
        FROM facturas_lineas
        WHERE factura_id = ?
        ORDER BY id
    """, (factura_id,))

    lineas_factura = [dict(row) for row in cursor.fetchall()]

    if not historico:
        conn.close()
        return {
            "ok": True,
            "factura_id": factura_id,
            "proveedor_nit": proveedor_nit,
            "proveedor_nombre": factura.get("proveedor_nombre"),
            "numero_factura": factura.get("numero_factura"),
            "requiere_revision": True,
            "confianza": 0,
                "mensaje": "No existe historico contable para este proveedor.",
            "propuesta": []
        }

    cuenta_gasto = None
    cuenta_iva = None
    cuenta_pagar = None
    centro_costo = None

    for h in historico:
        cuenta = h.get("cuenta_contable")
        nombre_cuenta = (h.get("nombre_cuenta") or "").lower()

        if not centro_costo and h.get("centro_costo"):
            centro_costo = h.get("centro_costo")

        if not cuenta_gasto:
            debito = float(h.get("debito") or 0)
            if debito > 0 and "iva" not in nombre_cuenta and cuenta != "240805":
                cuenta_gasto = h

        if not cuenta_iva:
            if "iva" in nombre_cuenta or cuenta == "240805":
                cuenta_iva = h

        if not cuenta_pagar:
            credito = float(h.get("credito") or 0)
            if credito > 0:
                cuenta_pagar = h

    if lineas_factura:
        if clasificacion.get("requiere_revision_concepto"):
            conn.close()
            return {
                "ok": True,
                "factura_id": factura_id,
                "proveedor_nit": proveedor_nit,
                "proveedor_nombre": factura.get("proveedor_nombre"),
                "numero_factura": factura.get("numero_factura"),
                "fecha_factura": factura.get("fecha_factura"),
                "subtotal": subtotal,
                "iva": iva,
                "total_pagar": total_pagar,
                "valor_gasto": valor_gasto,
                "total_historicos": len(historico),
                "confianza": 0,
                "requiere_revision": True,
                "estado": "REQUIERE_REVISION_CONCEPTO",
                "mensaje": "Una o mas lineas de la factura no pudieron clasificarse por producto/servicio.",
                "clasificacion": clasificacion,
                "propuesta": []
            }

        grupos = {}

        for linea in lineas_factura:
            concepto = linea.get("concepto_servicio")

            if not concepto:
                continue

            if concepto not in grupos:
                grupos[concepto] = {
                    "concepto_servicio": concepto,
                    "valor": 0,
                    "descripciones": [],
                    "confianzas": []
                }

            grupos[concepto]["valor"] += float(linea.get("valor_linea") or 0)
            grupos[concepto]["descripciones"].append(linea.get("descripcion"))
            grupos[concepto]["confianzas"].append(float(linea.get("clasificacion_confianza") or 0))

        if grupos and valor_gasto > 0:
            suma_grupos = round(sum(g["valor"] for g in grupos.values()), 2)
            diferencia = round(valor_gasto - suma_grupos, 2)

            if diferencia and abs(diferencia) <= max(1, round(valor_gasto * 0.02, 2)):
                primer_grupo = next(iter(grupos.values()))
                primer_grupo["valor"] = round(primer_grupo["valor"] + diferencia, 2)

        faltantes_mapeo = []
        propuesta = []

        for concepto, grupo in grupos.items():
            mapeo = buscar_mapeo_erp(cursor, proveedor_nit, concepto)

            if not mapeo:
                faltantes_mapeo.append({
                    "concepto_servicio": concepto,
                    "motivo": "No existe mapeo ERP activo para proveedor + concepto.",
                    "valor": round(grupo["valor"], 2),
                    "descripciones": grupo["descripciones"]
                })
                continue

            if not mapeo.get("item_type_erp") or not mapeo.get("item_code_erp"):
                faltantes_mapeo.append({
                    "concepto_servicio": concepto,
                    "mapeo_erp_id": mapeo.get("id"),
                    "motivo": "El mapeo ERP no tiene item_type_erp o item_code_erp.",
                    "valor": round(grupo["valor"], 2),
                    "descripciones": grupo["descripciones"]
                })
                continue

            propuesta.append({
                "tipo": "GASTO",
                "concepto_servicio": concepto,
                "cuenta_contable": mapeo.get("cuenta_contable"),
                "nombre_cuenta": mapeo.get("nombre_cuenta"),
                "centro_costo": centro_costo,
                "debito": round(grupo["valor"], 2),
                "credito": 0,
                "descripcion": " / ".join(d for d in grupo["descripciones"] if d)[:500],
                "mapeo_erp_id": mapeo.get("id"),
                "item_type_erp": mapeo.get("item_type_erp"),
                "item_code_erp": mapeo.get("item_code_erp"),
                "item_description_erp": mapeo.get("item_description_erp"),
                "tax_id_erp": mapeo.get("tax_id_erp"),
                "document_id_erp": mapeo.get("document_id_erp"),
                "payment_id_erp": mapeo.get("payment_id_erp")
            })

        if faltantes_mapeo:
            conn.close()
            return {
                "ok": True,
                "factura_id": factura_id,
                "proveedor_nit": proveedor_nit,
                "proveedor_nombre": factura.get("proveedor_nombre"),
                "numero_factura": factura.get("numero_factura"),
                "fecha_factura": factura.get("fecha_factura"),
                "subtotal": subtotal,
                "iva": iva,
                "total_pagar": total_pagar,
                "valor_gasto": valor_gasto,
                "total_historicos": len(historico),
                "confianza": 0,
                "requiere_revision": True,
                "estado": "REQUIERE_MAPEO_ERP",
                "mensaje": "Una o mas lineas clasificadas no tienen mapeo ERP suficiente.",
                "faltantes_mapeo": faltantes_mapeo,
                "clasificacion": clasificacion,
                "propuesta": []
            }

        if iva > 0:
            if not cuenta_iva:
                conn.close()
                return {
                    "ok": True,
                    "factura_id": factura_id,
                    "proveedor_nit": proveedor_nit,
                    "proveedor_nombre": factura.get("proveedor_nombre"),
                    "numero_factura": factura.get("numero_factura"),
                    "requiere_revision": True,
                    "estado": "REQUIERE_REVISION",
                    "confianza": 0,
                    "mensaje": "La factura tiene IVA, pero no se encontrÃƒÂ³ cuenta IVA en el histÃƒÂ³rico.",
                    "propuesta": []
                }

            propuesta.append({
                "tipo": "IVA",
                "concepto_servicio": "IVA",
                "cuenta_contable": cuenta_iva.get("cuenta_contable"),
                "nombre_cuenta": cuenta_iva.get("nombre_cuenta"),
                "centro_costo": centro_costo,
                "debito": iva,
                "credito": 0,
                "descripcion": "IVA descontable segÃƒÂºn histÃƒÂ³rico del proveedor"
            })

        if not cuenta_pagar:
            conn.close()
            return {
                "ok": True,
                "factura_id": factura_id,
                "proveedor_nit": proveedor_nit,
                "proveedor_nombre": factura.get("proveedor_nombre"),
                "numero_factura": factura.get("numero_factura"),
                "requiere_revision": True,
                "estado": "REQUIERE_REVISION",
                "confianza": 0,
                "mensaje": "No se encontrÃƒÂ³ cuenta por pagar en el histÃƒÂ³rico del proveedor.",
                "propuesta": []
            }

        propuesta.append({
            "tipo": "CXP",
            "concepto_servicio": "CXP",
            "cuenta_contable": cuenta_pagar.get("cuenta_contable"),
            "nombre_cuenta": cuenta_pagar.get("nombre_cuenta"),
            "centro_costo": centro_costo,
            "debito": 0,
            "credito": total_pagar,
            "descripcion": "Cuenta por pagar segÃƒÂºn histÃƒÂ³rico del proveedor"
        })

        confianza_lineas = [
            c.get("clasificacion_confianza") or 0
            for c in clasificacion.get("clasificaciones", [])
        ]
        confianza = min(confianza_lineas) if confianza_lineas else 80
        conn.close()

        return {
            "ok": True,
            "factura_id": factura_id,
            "proveedor_nit": proveedor_nit,
            "proveedor_nombre": factura.get("proveedor_nombre"),
            "numero_factura": factura.get("numero_factura"),
            "fecha_factura": factura.get("fecha_factura"),
            "subtotal": subtotal,
            "iva": iva,
            "total_pagar": total_pagar,
            "valor_gasto": valor_gasto,
            "total_historicos": len(historico),
            "confianza": confianza,
            "requiere_revision": False,
            "estado": "PROPUESTA_CONCEPTO_SERVICIO",
            "clasificacion": clasificacion,
            "propuesta": propuesta
        }

    propuesta = []

    if cuenta_gasto:
        propuesta.append({
            "tipo": "GASTO",
            "concepto_servicio": inferir_concepto_servicio(
                cuenta_gasto.get("nombre_cuenta"),
                cuenta_gasto.get("descripcion")
            ) or "OTRO",
            "cuenta_contable": cuenta_gasto.get("cuenta_contable"),
            "nombre_cuenta": cuenta_gasto.get("nombre_cuenta"),
            "centro_costo": centro_costo,
            "debito": valor_gasto,
            "credito": 0,
            "descripcion": "Gasto segÃºn histÃ³rico del proveedor"
        })

    if iva > 0 and cuenta_iva:
        propuesta.append({
            "tipo": "IVA",
            "concepto_servicio": "IVA",
            "cuenta_contable": cuenta_iva.get("cuenta_contable"),
            "nombre_cuenta": cuenta_iva.get("nombre_cuenta"),
            "centro_costo": centro_costo,
            "debito": iva,
            "credito": 0,
            "descripcion": "IVA descontable segÃºn histÃ³rico del proveedor"
        })

    if cuenta_pagar:
        propuesta.append({
            "tipo": "CXP",
            "concepto_servicio": "CXP",
            "cuenta_contable": cuenta_pagar.get("cuenta_contable"),
            "nombre_cuenta": cuenta_pagar.get("nombre_cuenta"),
            "centro_costo": centro_costo,
            "debito": 0,
            "credito": total_pagar,
            "descripcion": "Cuenta por pagar segÃºn histÃ³rico del proveedor"
        })

    requiere_revision = len(propuesta) == 0
    conn.close()

    return {
        "ok": True,
        "factura_id": factura_id,
        "proveedor_nit": proveedor_nit,
        "proveedor_nombre": factura.get("proveedor_nombre"),
        "numero_factura": factura.get("numero_factura"),
        "fecha_factura": factura.get("fecha_factura"),
        "subtotal": subtotal,
        "iva": iva,
        "total_pagar": total_pagar,
        "valor_gasto": valor_gasto,
        "total_historicos": len(historico),
        "confianza": 80 if propuesta else 0,
        "requiere_revision": requiere_revision,
        "estado": "PROPUESTA_HISTORICO_SIMPLE" if propuesta else "REQUIERE_REVISION",
        "propuesta": propuesta
    }


def limpiar_cuerpo_correo(texto_correo: str):
    if not texto_correo:
        return ""

    texto_limpio = unescape(texto_correo)

    texto_limpio = re.sub(r"<br\s*/?>", "\n", texto_limpio, flags=re.IGNORECASE)
    texto_limpio = re.sub(r"</p>", "\n", texto_limpio, flags=re.IGNORECASE)
    texto_limpio = re.sub(r"<[^>]+>", "", texto_limpio)

    return texto_limpio.strip()


def extraer_factura_id(subject: str, body: str):
    texto_busqueda = f"{subject}\n{body}"

    patrones = [
        r"FACTURA_ID\s*=\s*(\d+)",
        r"Factura ID\s+(\d+)",
        r"factura_id\s*:\s*(\d+)"
    ]

    for patron in patrones:
        match = re.search(patron, texto_busqueda, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))

    return None


def normalizar_tipo_linea_respuesta(tipo):
    tipo_normalizado = normalizar_clave_excel(tipo).upper()

    if tipo_normalizado == "CUENTA_POR_PAGAR":
        return "CXP"

    if tipo_normalizado in {"IMPUESTO_CONSUMO", "IMPOCONSUMO", "PROPINA", "AJUSTE", "OTRO_DEBITO"}:
        return "AJUSTE"

    if tipo_normalizado in ["GASTO", "IVA", "CXP"]:
        return tipo_normalizado

    return None


def construir_linea_respuesta(tipo, cuenta, nombre_cuenta, concepto_servicio, centro_costo_default):
    tipo_normalizado = normalizar_tipo_linea_respuesta(tipo)

    if not tipo_normalizado:
        raise ValueError(
            f"Tipo de linea no soportado: {tipo}. Use GASTO, IVA, AJUSTE o CXP."
        )

    cuenta_texto = str(cuenta or "").strip()
    nombre_texto = str(nombre_cuenta or "").strip()
    concepto_texto = str(concepto_servicio or "").strip() or None

    if not cuenta_texto:
        raise ValueError(
            "Cada linea contable debe incluir tipo y cuenta."
        )

    if not nombre_texto:
        if concepto_texto:
            nombre_texto = concepto_texto.replace("_", " ").title()
        elif tipo_normalizado == "IVA":
            nombre_texto = "IVA descontable"
        elif tipo_normalizado == "CXP":
            nombre_texto = "Cuenta por pagar"
        else:
            nombre_texto = "Gasto"

    return {
        "tipo": tipo_normalizado,
        "cuenta_contable": cuenta_texto,
        "nombre_cuenta": nombre_texto,
        "centro_costo": centro_costo_default,
        "concepto_servicio": concepto_texto,
        "descripcion": nombre_texto
    }


def parsear_linea_tabla_respuesta(line):
    if ";" in line:
        partes = [p.strip() for p in line.split(";")]
    elif "|" in line:
        partes = [p.strip() for p in line.split("|")]
    elif "\t" in line:
        partes = [p.strip() for p in line.split("\t")]
    else:
        partes = [p.strip() for p in re.split(r"\s{2,}", line)]

    if len(partes) < 3:
        return None

    tipo = partes[0]

    if not normalizar_tipo_linea_respuesta(tipo):
        return None

    cuenta = partes[1]

    if len(partes) == 3:
        nombre_cuenta = None
        concepto_servicio = partes[2]
    else:
        concepto_servicio = partes[-1]
        nombre_cuenta = " ".join(partes[2:-1]).strip()

    if not cuenta or not concepto_servicio:
        return None

    return tipo, cuenta, nombre_cuenta, concepto_servicio


def es_encabezado_tabla_respuesta(line):
    texto = normalizar_clave_excel(line)
    return (
        "tipo" in texto
        and "cuenta" in texto
        and "concepto" in texto
    )


def parsear_respuesta_causacion(subject: str, body: str):
    texto_limpio = limpiar_cuerpo_correo(body)

    if "#CAUSAR_FACTURA" not in texto_limpio:
        raise ValueError("No se encontro el marcador #CAUSAR_FACTURA en la respuesta.")

    factura_id = extraer_factura_id(subject, texto_limpio)

    centro_costo_default = None
    lineas = []

    for raw_line in texto_limpio.splitlines():
        line = raw_line.strip()

        if not line:
            continue

        if line.startswith(">"):
            continue

        if line.upper().startswith("CENTRO_COSTO="):
            centro_costo_default = line.split("=", 1)[1].strip()
            continue

        if line.upper().startswith("LINEA="):
            contenido = line.split("=", 1)[1].strip()
            separador = ";" if ";" in contenido else "|"
            partes = [p.strip() for p in contenido.split(separador)]

            if len(partes) < 3:
                raise ValueError(
                    f"Linea contable invalida. Debe tener al menos 3 partes: {line}"
                )

            lineas.append(construir_linea_respuesta(
                partes[0],
                partes[1],
                partes[2] if len(partes) >= 4 else None,
                partes[3] if len(partes) >= 4 else partes[2],
                centro_costo_default
            ))
            continue

        if es_encabezado_tabla_respuesta(line):
            continue

        linea_tabla = parsear_linea_tabla_respuesta(line)

        if linea_tabla:
            tipo, cuenta, nombre_cuenta, concepto_servicio = linea_tabla
            lineas.append(construir_linea_respuesta(
                tipo,
                cuenta,
                nombre_cuenta,
                concepto_servicio,
                centro_costo_default
            ))

    if not factura_id:
        raise ValueError("No se encontro FACTURA_ID en la respuesta.")

    if not lineas:
        raise ValueError("No se encontraron lineas contables en la respuesta.")

    return {
        "factura_id": factura_id,
        "centro_costo": centro_costo_default,
        "lineas": lineas
    }


def parsear_respuesta_proveedor_siigo(subject: str, body: str):
    texto_limpio = limpiar_cuerpo_correo(body)

    if "#CREAR_PROVEEDOR_SIIGO" not in texto_limpio:
        raise ValueError("No se encontro el marcador #CREAR_PROVEEDOR_SIIGO en la respuesta.")

    datos = {}

    for raw_line in texto_limpio.splitlines():
        line = raw_line.strip()

        if not line or line.startswith(">"):
            continue

        if "=" not in line:
            continue

        clave, valor = line.split("=", 1)
        clave = normalizar_clave_excel(clave)
        valor = valor.strip()

        if clave:
            datos[clave] = valor

    causacion_id = datos.get("causacion_id")

    if not causacion_id:
        match = re.search(r"CAUSACI[OÓ]N\s+(\d+)", subject or "", flags=re.IGNORECASE)
        causacion_id = match.group(1) if match else None

    if not causacion_id:
        raise ValueError("No se encontro CAUSACION_ID en la respuesta de proveedor.")

    try:
        causacion_id = int(causacion_id)
    except Exception:
        raise ValueError("CAUSACION_ID debe ser numerico.")

    return {
        "causacion_id": causacion_id,
        "extras": {
            "nombre": datos.get("nombre"),
            "nit": datos.get("nit"),
            "direccion": datos.get("direccion"),
            "departamento_codigo": datos.get("departamento_codigo"),
            "ciudad_codigo": datos.get("ciudad_codigo"),
            "pais_codigo": datos.get("pais_codigo") or "CO",
            "responsabilidad_fiscal": datos.get("responsabilidad_fiscal") or "R-99-PN",
            "telefono": datos.get("telefono"),
            "email": datos.get("email"),
            "check_digit": datos.get("digito_verificacion") or datos.get("check_digit")
        }
    }


def procesar_respuesta_proveedor_siigo(payload: dict):
    from siigo_client import SiigoClient

    subject = payload.get("subject", "")
    body = payload.get("body", "")
    datos = parsear_respuesta_proveedor_siigo(subject, body)
    causacion_id = datos["causacion_id"]

    resultado = construir_payload_siigo_compra_desde_causacion(causacion_id)
    factura = resultado["factura"]

    client = SiigoClient()
    proveedor = asegurar_proveedor_siigo(
        client,
        factura,
        causacion_id,
        datos.get("extras")
    )

    if proveedor.get("requiere_datos"):
        return respuesta_requiere_datos_proveedor(
            causacion_id,
            factura,
            proveedor.get("datos") or {}
        )

    return {
        "ok": True,
        "estado": "PROVEEDOR_SIIGO_CREADO",
        "mensaje": "Proveedor validado o creado en SIIGO. La compra puede reenviarse.",
        "factura_id": factura.get("id"),
        "causacion_id": causacion_id,
        "proveedor_nit": factura.get("proveedor_nit"),
        "proveedor_nombre": factura.get("proveedor_nombre"),
        "numero_factura": factura.get("numero_factura"),
        "proveedor_siigo": proveedor.get("proveedor_siigo"),
        "payload_proveedor": proveedor.get("payload_proveedor")
    }


def construir_lineas_respuesta_con_valores_factura(factura: dict, lineas_usuario: list):
    total_pagar = round(float(factura.get("total_pagar") or 0), 2)
    iva = round(float(factura.get("iva") or 0), 2)

    subtotal = round(float(factura.get("subtotal") or 0), 2)
    total_sin_impuestos = round(float(factura.get("total_sin_impuestos") or 0), 2)

    if total_pagar > 0:
        valor_gasto = round(total_pagar - iva, 2)
    elif subtotal > 0:
        valor_gasto = subtotal
    elif total_sin_impuestos > 0:
        valor_gasto = total_sin_impuestos
    else:
        valor_gasto = 0

    cuentas = {}

    for linea in lineas_usuario:
        tipo = linea.get("tipo")

        cuentas[tipo] = linea

    if "GASTO" not in cuentas:
        raise ValueError("La respuesta debe incluir una lÃ­nea tipo GASTO.")

    if "CXP" not in cuentas:
        raise ValueError("La respuesta debe incluir una lÃ­nea tipo CXP.")

    if iva > 0 and "IVA" not in cuentas:
        raise ValueError("La factura tiene IVA, por lo tanto debe incluir una lÃ­nea tipo IVA.")

    lineas_calculadas = []

    gasto = cuentas["GASTO"]

    lineas_calculadas.append({
        "tipo": "GASTO",
        "cuenta_contable": gasto.get("cuenta_contable"),
        "nombre_cuenta": gasto.get("nombre_cuenta"),
        "centro_costo": gasto.get("centro_costo"),
        "debito": valor_gasto,
        "credito": 0,
        "concepto_servicio": gasto.get("concepto_servicio") or inferir_concepto_servicio(
            gasto.get("nombre_cuenta"),
            gasto.get("descripcion")
        ) or "OTRO",
        "descripcion": gasto.get("descripcion")
    })

    if iva > 0:
        iva_linea = cuentas["IVA"]

        lineas_calculadas.append({
            "tipo": "IVA",
            "cuenta_contable": iva_linea.get("cuenta_contable"),
            "nombre_cuenta": iva_linea.get("nombre_cuenta"),
            "centro_costo": iva_linea.get("centro_costo"),
            "debito": iva,
            "credito": 0,
            "concepto_servicio": iva_linea.get("concepto_servicio") or "IVA",
            "descripcion": iva_linea.get("descripcion")
        })

    cxp = cuentas["CXP"]

    lineas_calculadas.append({
        "tipo": "CXP",
        "cuenta_contable": cxp.get("cuenta_contable"),
        "nombre_cuenta": cxp.get("nombre_cuenta"),
        "centro_costo": cxp.get("centro_costo"),
        "debito": 0,
        "credito": total_pagar,
        "concepto_servicio": cxp.get("concepto_servicio") or "CXP",
        "descripcion": cxp.get("descripcion")
    })

    total_debito = round(sum(float(l.get("debito") or 0) for l in lineas_calculadas), 2)
    total_credito = round(sum(float(l.get("credito") or 0) for l in lineas_calculadas), 2)

    if total_debito != total_credito:
        raise ValueError(
            f"La contabilizaciÃ³n calculada con valores de la factura estÃ¡ descuadrada. "
            f"DÃ©bito={total_debito}, CrÃ©dito={total_credito}"
        )

    return {
        "lineas": lineas_calculadas,
        "total_debito": total_debito,
        "total_credito": total_credito,
        "total_pagar_factura": total_pagar,
        "iva_factura": iva,
        "valor_gasto_factura": valor_gasto
    }


def upsert_mapeo_erp_desde_linea_manual(cursor, factura: dict, linea: dict):
    if linea.get("tipo") not in {"GASTO", "AJUSTE"}:
        return

    proveedor_nit = factura.get("proveedor_nit")
    concepto = linea.get("concepto_servicio")
    cuenta = linea.get("cuenta_contable")

    if not proveedor_nit or not concepto or not cuenta:
        return

    cursor.execute("""
        SELECT id
        FROM mapeo_erp
        WHERE activo = 1
          AND erp = 'SIIGO'
          AND proveedor_nit = ?
          AND concepto_servicio = ?
          AND cuenta_contable = ?
        LIMIT 1
    """, (proveedor_nit, concepto, cuenta))

    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE mapeo_erp
            SET
                proveedor_nombre = ?,
                nombre_cuenta = ?,
                item_type_erp = COALESCE(NULLIF(item_type_erp, ''), 'Account'),
                item_code_erp = COALESCE(NULLIF(item_code_erp, ''), ?),
                item_description_erp = COALESCE(NULLIF(item_description_erp, ''), ?),
                observacion = COALESCE(NULLIF(observacion, ''), 'Aprendido desde respuesta manual')
            WHERE id = ?
        """, (
            factura.get("proveedor_nombre"),
            linea.get("nombre_cuenta"),
            cuenta,
            linea.get("nombre_cuenta"),
            existente[0]
        ))
        return

    cursor.execute("""
        INSERT INTO mapeo_erp (
            erp,
            proveedor_nit,
            proveedor_nombre,
            concepto_servicio,
            cuenta_contable,
            nombre_cuenta,
            item_type_erp,
            item_code_erp,
            item_description_erp,
            activo,
            observacion
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        "SIIGO",
        proveedor_nit,
        factura.get("proveedor_nombre"),
        concepto,
        cuenta,
        linea.get("nombre_cuenta"),
        "Account",
        cuenta,
        linea.get("nombre_cuenta"),
        1,
        "Aprendido desde respuesta manual"
    ))


def upsert_reglas_aprendidas_desde_respuesta_manual(cursor, factura: dict, lineas_factura: list, lineas_contables: list):
    proveedor_nit = factura.get("proveedor_nit")

    if not proveedor_nit:
        return []

    lineas_gasto = [
        linea
        for linea in lineas_contables
        if linea.get("tipo") in {"GASTO", "AJUSTE"}
        and linea.get("concepto_servicio")
        and linea.get("cuenta_contable")
    ]

    if not lineas_gasto:
        return []

    if len(lineas_gasto) == 1:
        linea_base = lineas_gasto[0]
    else:
        linea_base = sorted(
            lineas_gasto,
            key=lambda linea: float(linea.get("debito") or 0),
            reverse=True
        )[0]

    reglas = []

    for linea_factura in lineas_factura:
        descripcion = linea_factura.get("descripcion")
        palabras = generar_palabras_clave_descripcion_aprendida(descripcion)

        if not palabras:
            continue

        palabras_clave = ", ".join(palabras)
        concepto = linea_base.get("concepto_servicio")
        cuenta = linea_base.get("cuenta_contable")
        nombre_cuenta = linea_base.get("nombre_cuenta")

        cursor.execute("""
            SELECT id
            FROM reglas_concepto_servicio
            WHERE erp = 'SIIGO'
              AND proveedor_nit = ?
              AND concepto_servicio = ?
              AND palabras_clave = ?
            LIMIT 1
        """, (proveedor_nit, concepto, palabras_clave))

        existente = cursor.fetchone()

        if existente:
            regla_id = existente[0]
            cursor.execute("""
                UPDATE reglas_concepto_servicio
                SET
                    cuenta_contable = ?,
                    nombre_cuenta = ?,
                    item_type_erp = 'Account',
                    item_code_erp = ?,
                    prioridad = 20,
                    activo = 1
                WHERE id = ?
            """, (
                cuenta,
                nombre_cuenta,
                cuenta,
                regla_id
            ))
            accion = "actualizada"
        else:
            cursor.execute("""
                INSERT INTO reglas_concepto_servicio (
                    erp,
                    proveedor_nit,
                    concepto_servicio,
                    palabras_clave,
                    cuenta_contable,
                    nombre_cuenta,
                    item_type_erp,
                    item_code_erp,
                    prioridad,
                    activo
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                "SIIGO",
                proveedor_nit,
                concepto,
                palabras_clave,
                cuenta,
                nombre_cuenta,
                "Account",
                cuenta,
                20,
                1
            ))
            regla_id = cursor.lastrowid
            accion = "creada"

        cursor.execute("""
            UPDATE facturas_lineas
            SET
                concepto_servicio = ?,
                clasificacion_fuente = 'respuesta_manual_aprendida',
                clasificacion_confianza = 100
            WHERE id = ?
        """, (
            concepto,
            linea_factura.get("id")
        ))

        reglas.append({
            "regla_id": regla_id,
            "accion": accion,
            "linea_factura_id": linea_factura.get("id"),
            "descripcion": descripcion,
            "concepto_servicio": concepto,
            "cuenta_contable": cuenta,
            "palabras_clave": palabras_clave
        })

    return reglas


@app.get("/health")
def health():
    return {
        "status": "ok",
        "message": "API de Facturas IA funcionando correctamente"
    }


@app.get("/facturas/listar")
def listar_facturas():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            numero_factura,
            fecha_factura,
            cufe,
            proveedor_nombre,
            proveedor_nit,
            cliente_nombre,
            cliente_nit,
            subtotal,
            iva,
            total_pagar,
            fecha_procesamiento
        FROM facturas_recibidas
        ORDER BY id DESC
        LIMIT 50
    """)

    rows = cursor.fetchall()
    conn.close()

    return {
        "ok": True,
        "facturas": [dict(row) for row in rows]
    }


@app.get("/clasificaciones/conceptos")
def listar_conceptos_clasificacion():
    conceptos_base = [
        {
            "concepto": concepto,
            "palabras_clave": palabras
        }
        for concepto, palabras in CONCEPTOS_SERVICIO_KEYWORDS
    ]

    conceptos = {
        item["concepto"]
        for item in conceptos_base
        if item.get("concepto")
    }

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            erp,
            proveedor_nit,
            concepto_servicio,
            palabras_clave,
            cuenta_contable,
            nombre_cuenta,
            prioridad,
            activo
        FROM reglas_concepto_servicio
        WHERE activo = 1
        ORDER BY prioridad ASC, concepto_servicio ASC, id ASC
    """)
    reglas = [dict(row) for row in cursor.fetchall()]

    for regla in reglas:
        if regla.get("concepto_servicio"):
            conceptos.add(regla.get("concepto_servicio"))

    cursor.execute("""
        SELECT
            concepto_servicio,
            COUNT(*) AS total_mapeos
        FROM mapeo_erp
        WHERE activo = 1
          AND concepto_servicio IS NOT NULL
          AND TRIM(concepto_servicio) <> ''
        GROUP BY concepto_servicio
        ORDER BY concepto_servicio ASC
    """)
    mapeos = [dict(row) for row in cursor.fetchall()]

    for mapeo in mapeos:
        if mapeo.get("concepto_servicio"):
            conceptos.add(mapeo.get("concepto_servicio"))

    conn.close()

    conceptos_ordenados = sorted(conceptos)

    return {
        "ok": True,
        "total_conceptos": len(conceptos_ordenados),
        "conceptos": conceptos_ordenados,
        "conceptos_base": conceptos_base,
        "reglas_activas": reglas,
        "mapeos_por_concepto": mapeos
    }


@app.post("/clasificaciones/mapeo-erp")
def guardar_mapeo_erp_clasificacion(payload: dict = Body(...)):
    proveedor_nit = valor_texto_excel(payload.get("proveedor_nit"))
    proveedor_nombre = valor_texto_excel(payload.get("proveedor_nombre"))
    concepto_servicio = valor_texto_excel(payload.get("concepto_servicio"))
    cuenta_contable = valor_texto_excel(payload.get("cuenta_contable"))
    nombre_cuenta = valor_texto_excel(payload.get("nombre_cuenta"))
    item_type_erp = valor_texto_excel(payload.get("item_type_erp")) or "Account"
    item_code_erp = valor_texto_excel(payload.get("item_code_erp")) or cuenta_contable
    item_description_erp = valor_texto_excel(payload.get("item_description_erp")) or nombre_cuenta
    palabras_clave = valor_texto_excel(payload.get("palabras_clave"))
    activo = valor_activo_excel(payload.get("activo", True))
    prioridad = int(valor_numero_excel(payload.get("prioridad"), 20))

    if concepto_servicio:
        concepto_servicio = concepto_servicio.strip().upper().replace(" ", "_")

    if not proveedor_nit or not concepto_servicio:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "DATOS_INCOMPLETOS",
                "mensaje": "proveedor_nit y concepto_servicio son obligatorios."
            }
        )

    if not cuenta_contable and not item_code_erp:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "DATOS_INCOMPLETOS",
                "mensaje": "Debe enviar cuenta_contable o item_code_erp."
            }
        )

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM mapeo_erp
        WHERE proveedor_nit = ?
          AND concepto_servicio = ?
        ORDER BY activo DESC, id ASC
        LIMIT 1
    """, (proveedor_nit, concepto_servicio))

    existente = cursor.fetchone()

    if existente:
        mapeo_id = existente["id"]
        cursor.execute("""
            UPDATE mapeo_erp
            SET
                erp = ?,
                proveedor_nombre = ?,
                cuenta_contable = ?,
                nombre_cuenta = ?,
                item_type_erp = ?,
                item_code_erp = ?,
                item_description_erp = ?,
                document_id_erp = ?,
                payment_id_erp = ?,
                tax_id_erp = ?,
                activo = ?,
                observacion = ?
            WHERE id = ?
        """, (
            valor_texto_excel(payload.get("erp")) or "SIIGO",
            proveedor_nombre,
            cuenta_contable,
            nombre_cuenta,
            item_type_erp,
            item_code_erp,
            item_description_erp,
            valor_texto_excel(payload.get("document_id_erp")),
            valor_texto_excel(payload.get("payment_id_erp")),
            valor_texto_excel(payload.get("tax_id_erp")),
            activo,
            valor_texto_excel(payload.get("observacion")) or "Actualizado desde Swagger",
            mapeo_id
        ))
        accion = "actualizado"
    else:
        cursor.execute("""
            INSERT INTO mapeo_erp (
                erp,
                proveedor_nit,
                proveedor_nombre,
                concepto_servicio,
                cuenta_contable,
                nombre_cuenta,
                item_type_erp,
                item_code_erp,
                item_description_erp,
                document_id_erp,
                payment_id_erp,
                tax_id_erp,
                activo,
                observacion
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            valor_texto_excel(payload.get("erp")) or "SIIGO",
            proveedor_nit,
            proveedor_nombre,
            concepto_servicio,
            cuenta_contable,
            nombre_cuenta,
            item_type_erp,
            item_code_erp,
            item_description_erp,
            valor_texto_excel(payload.get("document_id_erp")),
            valor_texto_excel(payload.get("payment_id_erp")),
            valor_texto_excel(payload.get("tax_id_erp")),
            activo,
            valor_texto_excel(payload.get("observacion")) or "Creado desde Swagger"
        ))
        mapeo_id = cursor.lastrowid
        accion = "creado"

    regla_id = None

    if palabras_clave:
        cursor.execute("""
            SELECT id
            FROM reglas_concepto_servicio
            WHERE proveedor_nit = ?
              AND concepto_servicio = ?
              AND palabras_clave = ?
            LIMIT 1
        """, (proveedor_nit, concepto_servicio, palabras_clave))

        regla_existente = cursor.fetchone()

        if regla_existente:
            regla_id = regla_existente["id"]
            cursor.execute("""
                UPDATE reglas_concepto_servicio
                SET
                    cuenta_contable = ?,
                    nombre_cuenta = ?,
                    item_type_erp = ?,
                    item_code_erp = ?,
                    tax_id_erp = ?,
                    prioridad = ?,
                    activo = 1
                WHERE id = ?
            """, (
                cuenta_contable,
                nombre_cuenta,
                item_type_erp,
                item_code_erp,
                valor_texto_excel(payload.get("tax_id_erp")),
                prioridad,
                regla_id
            ))
        else:
            cursor.execute("""
                INSERT INTO reglas_concepto_servicio (
                    erp,
                    proveedor_nit,
                    concepto_servicio,
                    palabras_clave,
                    cuenta_contable,
                    nombre_cuenta,
                    item_type_erp,
                    item_code_erp,
                    tax_id_erp,
                    prioridad,
                    activo
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (
                valor_texto_excel(payload.get("erp")) or "SIIGO",
                proveedor_nit,
                concepto_servicio,
                palabras_clave,
                cuenta_contable,
                nombre_cuenta,
                item_type_erp,
                item_code_erp,
                valor_texto_excel(payload.get("tax_id_erp")),
                prioridad
            ))
            regla_id = cursor.lastrowid

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "accion": accion,
        "mapeo_erp_id": mapeo_id,
        "regla_id": regla_id,
        "proveedor_nit": proveedor_nit,
        "proveedor_nombre": proveedor_nombre,
        "concepto_servicio": concepto_servicio,
        "cuenta_contable": cuenta_contable,
        "item_type_erp": item_type_erp,
        "item_code_erp": item_code_erp,
        "palabras_clave": palabras_clave
    }


@app.get("/facturas/{factura_id}/detalle")
def obtener_detalle_factura(factura_id: int):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM facturas_recibidas
        WHERE id = ?
    """, (factura_id,))

    factura_row = cursor.fetchone()

    if not factura_row:
        conn.close()
        return JSONResponse(
            status_code=404,
            content={
                "ok": False,
                "estado": "FACTURA_NO_ENCONTRADA",
                "mensaje": f"No existe la factura {factura_id}.",
                "factura_id": factura_id
            }
        )

    factura = dict(factura_row)

    cursor.execute("""
        SELECT *
        FROM facturas_lineas
        WHERE factura_id = ?
        ORDER BY id
    """, (factura_id,))
    lineas = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
        SELECT *
        FROM causaciones
        WHERE factura_id = ?
        ORDER BY id DESC
        LIMIT 1
    """, (factura_id,))
    causacion_row = cursor.fetchone()

    causacion = dict(causacion_row) if causacion_row else None

    conn.close()

    return {
        "ok": True,
        "factura": factura,
        "lineas": lineas,
        "causacion": causacion,
        "tiene_pdf": bool(factura.get("pdf_principal") and os.path.exists(factura.get("pdf_principal")))
    }


@app.get("/facturas/{factura_id}/pdf")
def descargar_pdf_factura(factura_id: int):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            numero_factura,
            proveedor_nombre,
            pdf_principal
        FROM facturas_recibidas
        WHERE id = ?
    """, (factura_id,))

    factura_row = cursor.fetchone()
    conn.close()

    if not factura_row:
        return JSONResponse(
            status_code=404,
            content={
                "ok": False,
                "estado": "FACTURA_NO_ENCONTRADA",
                "mensaje": f"No existe la factura {factura_id}.",
                "factura_id": factura_id
            }
        )

    factura = dict(factura_row)
    pdf_path = factura.get("pdf_principal")

    if not pdf_path or not os.path.exists(pdf_path):
        return JSONResponse(
            status_code=404,
            content={
                "ok": False,
                "estado": "PDF_NO_ENCONTRADO",
                "mensaje": "La factura no tiene PDF asociado o el archivo no existe.",
                "factura_id": factura_id,
                "pdf_principal": pdf_path
            }
        )

    proveedor = re.sub(r"[^A-Za-z0-9._ -]+", "", factura.get("proveedor_nombre") or "Proveedor").strip()
    numero = re.sub(r"[^A-Za-z0-9._ -]+", "", factura.get("numero_factura") or str(factura_id)).strip()
    filename = f"{proveedor} - {numero}.pdf"

    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        filename=filename
    )


def obtener_env_requerido(nombre: str):
    valor = os.getenv(nombre)
    if not valor:
        raise ValueError(f"Variable de entorno requerida no configurada: {nombre}")
    return valor


def normalizar_message_id(message_id: str):
    texto = str(message_id or "").strip()
    if not texto:
        return ""
    texto = texto.strip("<>")
    return f"<{texto}>"


def mailbox_imap(mailbox: str):
    nombre = str(mailbox or "INBOX").strip() or "INBOX"
    if nombre.upper() == "INBOX":
        return "INBOX"
    return '"' + nombre.replace("\\", "\\\\").replace('"', '\\"') + '"'


def uid_desde_respuesta_search(data):
    if not data:
        return None

    primera = data[0]
    if isinstance(primera, bytes):
        texto = primera.decode("utf-8", errors="ignore").strip()
    else:
        texto = str(primera or "").strip()

    if not texto:
        return None

    return texto.split()[0]


def mailboxes_busqueda_imap(origen):
    candidatos = [
        origen,
        os.getenv("IMAP_SOURCE_MAILBOX", "INBOX"),
        "INBOX",
        "[Gmail]/All Mail",
        "[Gmail]/Todos",
        "[Gmail]/Todos los correos",
        "All Mail",
        "Todos"
    ]
    resultado = []

    for candidato in candidatos:
        candidato = str(candidato or "").strip()

        if candidato and candidato not in resultado:
            resultado.append(candidato)

    return resultado


def archivar_correo_imap(
    email_uid: str = "",
    email_message_id_header: str = "",
    source_mailbox: str = "INBOX",
    target_mailbox: str = "Facturas Procesadas",
):
    host = obtener_env_requerido("IMAP_HOST")
    user = obtener_env_requerido("IMAP_USER")
    password = obtener_env_requerido("IMAP_PASSWORD")
    port = int(os.getenv("IMAP_PORT", "993"))
    secure = os.getenv("IMAP_SECURE", "true").lower() != "false"

    origen = source_mailbox or os.getenv("IMAP_SOURCE_MAILBOX", "INBOX")
    destino = target_mailbox or os.getenv("IMAP_ARCHIVE_MAILBOX", "Facturas Procesadas")
    message_id = normalizar_message_id(email_message_id_header)
    uid = str(email_uid or "").strip()

    if not uid and not message_id:
        raise ValueError("Debe enviar email_uid o email_message_id_header para ubicar el correo.")

    imap = imaplib.IMAP4_SSL(host, port) if secure else imaplib.IMAP4(host, port)

    try:
        imap.login(user, password)

        mailbox_encontrado = origen

        if not uid:
            for mailbox in mailboxes_busqueda_imap(origen):
                status, _ = imap.select(mailbox_imap(mailbox), readonly=False)

                if status != "OK":
                    continue

                status, data = imap.uid("SEARCH", None, "HEADER", "Message-ID", message_id)

                if status == "OK":
                    uid = uid_desde_respuesta_search(data)

                if uid:
                    mailbox_encontrado = mailbox
                    break

        else:
            status, _ = imap.select(mailbox_imap(origen), readonly=False)
            if status != "OK":
                raise ValueError(f"No se pudo abrir el mailbox origen IMAP: {origen}")

        if not uid:
            raise ValueError(
                "No se encontro el correo en IMAP por UID ni Message-ID. "
                "Revise si n8n esta enviando email_uid o email_message_id_header."
            )

        imap.create(mailbox_imap(destino))

        move_status, move_data = imap.uid("MOVE", uid, mailbox_imap(destino))
        metodo = "UID MOVE"

        if move_status != "OK":
            copy_status, copy_data = imap.uid("COPY", uid, mailbox_imap(destino))
            if copy_status != "OK":
                detalle = copy_data or move_data
                raise ValueError(f"No se pudo mover/copiar el correo en IMAP: {detalle}")

            store_status, store_data = imap.uid("STORE", uid, "+FLAGS", r"(\Deleted)")
            if store_status != "OK":
                raise ValueError(f"El correo se copio pero no se pudo quitar de {origen}: {store_data}")

            imap.expunge()
            metodo = "COPY + DELETE"

        return {
            "ok": True,
            "estado": "CORREO_ARCHIVADO_IMAP",
            "mensaje": "Correo movido por IMAP correctamente.",
            "uid": uid,
            "source_mailbox": mailbox_encontrado,
            "target_mailbox": destino,
            "metodo": metodo
        }

    finally:
        try:
            imap.logout()
        except Exception:
            pass


@app.post("/correo/archivar-imap")
def archivar_correo_imap_endpoint(payload: dict = Body(...)):
    try:
        return archivar_correo_imap(
            email_uid=payload.get("email_uid", ""),
            email_message_id_header=payload.get("email_message_id_header", ""),
            source_mailbox=payload.get("source_mailbox") or os.getenv("IMAP_SOURCE_MAILBOX", "INBOX"),
            target_mailbox=payload.get("target_mailbox") or os.getenv("IMAP_ARCHIVE_MAILBOX", "Facturas Procesadas"),
        )
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "ERROR_ARCHIVANDO_IMAP",
                "mensaje": str(exc)
            }
        )


@app.post("/facturas/{factura_id}/clasificar-conceptos")
def clasificar_conceptos_factura(factura_id: int, guardar: bool = True):
    resultado = clasificar_lineas_factura(factura_id, guardar=guardar)

    if not resultado.get("ok"):
        return JSONResponse(
            status_code=resultado.get("status_code", 400),
            content={
                "ok": False,
                "error": resultado.get("error", "No fue posible clasificar la factura."),
                "factura_id": factura_id
            }
        )

    return resultado


@app.get("/proveedores/listar")
def listar_proveedores():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            nit,
            nombre,
            ciudad,
            regimen,
            responsabilidad_fiscal,
            cuenta_gasto_default,
            cuenta_iva_default,
            cuenta_retencion_default,
            activo,
            fecha_creacion
        FROM proveedores
        ORDER BY nombre
    """)

    rows = cursor.fetchall()
    conn.close()

    return {
        "ok": True,
        "proveedores": [dict(row) for row in rows]
    }


@app.get("/historico/proveedor/{nit}")
def buscar_historico_proveedor(nit: str):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            proveedor_nit,
            proveedor_nombre,
            descripcion,
            cuenta_contable,
            nombre_cuenta,
            centro_costo,
            debito,
            credito,
            cuenta_iva,
            cuenta_retencion,
            fuente,
            fecha_documento
        FROM contabilizaciones_historicas
        WHERE proveedor_nit = ?
        ORDER BY fecha_documento DESC, id DESC
        LIMIT 20
    """, (nit,))

    rows = cursor.fetchall()
    conn.close()

    return {
        "ok": True,
        "nit": nit,
        "total_registros": len(rows),
        "historico": [dict(row) for row in rows]
    }


@app.post("/historico/cargar-archivo")
async def cargar_archivo_historico(file: UploadFile = File(...), reemplazar: bool = False):
    nombre_archivo = file.filename or "historico.xlsx"
    extension = os.path.splitext(nombre_archivo)[1].lower()

    if extension not in {".xlsx", ".xlsm"}:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": "El archivo debe ser .xlsx o .xlsm.",
                "archivo": nombre_archivo
            }
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": "Falta instalar openpyxl. Ejecuta: python -m pip install -r .\\requirements.txt"
            }
        )

    contenido = await file.read()

    try:
        workbook = load_workbook(
            filename=BytesIO(contenido),
            read_only=True,
            data_only=True
        )
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": f"No se pudo leer el Excel: {exc}",
                "archivo": nombre_archivo
            }
        )

    hoja_historico = obtener_hoja_excel(
        workbook,
        "Historico_Contable",
        "HistÃ³rico Contable",
        "Historico Contable"
    )
    hoja_mapeo = obtener_hoja_excel(
        workbook,
        "Mapeo_ERP",
        "Mapeo ERP"
    )
    hoja_reglas = obtener_hoja_excel(
        workbook,
        "Reglas",
        "Reglas_Concepto_Servicio",
        "Reglas Concepto Servicio"
    )

    if hoja_historico is None and hoja_mapeo is None and hoja_reglas is None:
        hojas_disponibles = workbook.sheetnames
        workbook.close()
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": "El Excel no contiene hojas Historico_Contable, Mapeo_ERP ni Reglas.",
                "hojas_disponibles": hojas_disponibles
            }
        )

    filas_historico = leer_filas_excel(hoja_historico) if hoja_historico else []
    filas_mapeo = leer_filas_excel(hoja_mapeo) if hoja_mapeo else []
    filas_reglas = leer_filas_excel(hoja_reglas) if hoja_reglas else []

    historicos = [
        registro
        for registro in (
            construir_historico_desde_excel(fila, nombre_archivo)
            for fila in filas_historico
        )
        if registro
    ]
    mapeos = [
        registro
        for registro in (
            construir_mapeo_desde_excel(fila)
            for fila in filas_mapeo
        )
        if registro
    ]
    reglas = [
        registro
        for registro in (
            construir_regla_desde_excel(fila)
            for fila in filas_reglas
        )
        if registro
    ]

    workbook.close()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        if reemplazar:
            cursor.execute("DELETE FROM contabilizaciones_historicas")
            cursor.execute("DELETE FROM mapeo_erp")
            cursor.execute("DELETE FROM reglas_concepto_servicio")

        cursor.executemany("""
            INSERT INTO contabilizaciones_historicas (
                proveedor_nit,
                proveedor_nombre,
                descripcion,
                cuenta_contable,
                nombre_cuenta,
                centro_costo,
                debito,
                credito,
                cuenta_iva,
                cuenta_retencion,
                fuente,
                fecha_documento
            )
            VALUES (
                :proveedor_nit,
                :proveedor_nombre,
                :descripcion,
                :cuenta_contable,
                :nombre_cuenta,
                :centro_costo,
                :debito,
                :credito,
                :cuenta_iva,
                :cuenta_retencion,
                :fuente,
                :fecha_documento
            )
        """, historicos)

        cursor.executemany("""
            INSERT INTO mapeo_erp (
                erp,
                proveedor_nit,
                proveedor_nombre,
                concepto_servicio,
                cuenta_contable,
                nombre_cuenta,
                item_type_erp,
                item_code_erp,
                item_description_erp,
                document_id_erp,
                payment_id_erp,
                tax_id_erp,
                activo,
                observacion
            )
            VALUES (
                :erp,
                :proveedor_nit,
                :proveedor_nombre,
                :concepto_servicio,
                :cuenta_contable,
                :nombre_cuenta,
                :item_type_erp,
                :item_code_erp,
                :item_description_erp,
                :document_id_erp,
                :payment_id_erp,
                :tax_id_erp,
                :activo,
                :observacion
            )
        """, mapeos)

        cursor.executemany("""
            INSERT INTO reglas_concepto_servicio (
                erp,
                proveedor_nit,
                concepto_servicio,
                palabras_clave,
                cuenta_contable,
                nombre_cuenta,
                item_type_erp,
                item_code_erp,
                tax_id_erp,
                prioridad,
                activo
            )
            VALUES (
                :erp,
                :proveedor_nit,
                :concepto_servicio,
                :palabras_clave,
                :cuenta_contable,
                :nombre_cuenta,
                :item_type_erp,
                :item_code_erp,
                :tax_id_erp,
                :prioridad,
                :activo
            )
        """, reglas)

        conn.commit()

    except Exception as exc:
        conn.rollback()
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": f"No se pudo cargar el historico: {exc}",
                "archivo": nombre_archivo
            }
        )

    finally:
        conn.close()

    return {
        "ok": True,
        "archivo": nombre_archivo,
        "reemplazar": reemplazar,
        "hojas_procesadas": {
            "Historico_Contable": hoja_historico is not None,
            "Mapeo_ERP": hoja_mapeo is not None,
            "Reglas": hoja_reglas is not None
        },
        "filas_leidas": {
            "Historico_Contable": len(filas_historico),
            "Mapeo_ERP": len(filas_mapeo),
            "Reglas": len(filas_reglas)
        },
        "filas_insertadas": {
            "contabilizaciones_historicas": len(historicos),
            "mapeo_erp": len(mapeos),
            "reglas_concepto_servicio": len(reglas)
        },
        "filas_omitidas": {
            "Historico_Contable": len(filas_historico) - len(historicos),
            "Mapeo_ERP": len(filas_mapeo) - len(mapeos),
            "Reglas": len(filas_reglas) - len(reglas)
        }
    }


@app.get("/causacion/proponer/{factura_id}")
def proponer_causacion(factura_id: int):
    resultado = construir_propuesta_causacion(factura_id)

    if not resultado.get("ok") and resultado.get("status_code") == 404:
        return JSONResponse(
            status_code=404,
            content={
                "ok": False,
                "error": resultado.get("error")
            }
        )

    return resultado


@app.post("/causacion/causar/{factura_id}")
def causar_factura(factura_id: int):
    propuesta = construir_propuesta_causacion(factura_id)

    if not propuesta.get("ok"):
        return JSONResponse(
            status_code=propuesta.get("status_code", 400),
            content={
                "ok": False,
                "error": propuesta.get("error", "No fue posible construir la causaciÃ³n.")
            }
        )

    if propuesta.get("requiere_revision"):
        estado_revision = propuesta.get("estado", "REQUIERE_REVISION")
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": estado_revision,
                "mensaje": propuesta.get("mensaje", "La factura requiere revision antes de causar."),
                "factura_id": factura_id,
                "proveedor_nit": propuesta.get("proveedor_nit"),
                "proveedor_nombre": propuesta.get("proveedor_nombre"),
                "numero_factura": propuesta.get("numero_factura"),
                "fecha_factura": propuesta.get("fecha_factura"),
                "subtotal": propuesta.get("subtotal"),
                "iva": propuesta.get("iva"),
                "total_pagar": propuesta.get("total_pagar"),
                "valor_gasto": propuesta.get("valor_gasto"),
                "confianza": propuesta.get("confianza"),
                "total_historicos": propuesta.get("total_historicos"),
                "clasificacion": propuesta.get("clasificacion"),
                "faltantes_mapeo": propuesta.get("faltantes_mapeo"),
                "propuesta": propuesta.get("propuesta", [])
            }
        )

    lineas = propuesta.get("propuesta", [])

    if not lineas:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "SIN_LINEAS",
                "mensaje": "No se generaron lÃ­neas contables para causar.",
                "factura_id": factura_id
            }
        )

    total_debito = round(sum(float(l.get("debito") or 0) for l in lineas), 2)
    total_credito = round(sum(float(l.get("credito") or 0) for l in lineas), 2)

    if total_debito != total_credito:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "DESCUADRADA",
                "mensaje": "La causaciÃ³n no cuadra. DÃ©bitos y crÃ©ditos son diferentes.",
                "factura_id": factura_id,
                "total_debito": total_debito,
                "total_credito": total_credito,
                "diferencia": round(total_debito - total_credito, 2),
                "lineas": lineas
            }
        )

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM causaciones
        WHERE factura_id = ?
          AND estado IN ('CAUSADA_SIMULADA', 'CAUSADA_RESPUESTA_MANUAL')
        LIMIT 1
    """, (factura_id,))

    existente = cursor.fetchone()

    if existente:
        conn.close()
        return {
            "ok": True,
            "estado": "YA_CAUSADA",
            "mensaje": "La factura ya tenÃ­a una causaciÃ³n registrada.",
            "factura_id": factura_id,
            "causacion_id": existente[0],
            "total_debito": total_debito,
            "total_credito": total_credito,
            "lineas": lineas
        }

    cursor.execute("""
        INSERT INTO causaciones (
            factura_id,
            proveedor_nit,
            proveedor_nombre,
            numero_factura,
            estado,
            confianza,
            total_debito,
            total_credito,
            mensaje
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        factura_id,
        propuesta.get("proveedor_nit"),
        propuesta.get("proveedor_nombre"),
        propuesta.get("numero_factura"),
        "CAUSADA_SIMULADA",
        propuesta.get("confianza"),
        total_debito,
        total_credito,
        "Factura causada automÃ¡ticamente en modo simulaciÃ³n usando valores reales del XML."
    ))

    causacion_id = cursor.lastrowid

    for linea in lineas:
        cursor.execute("""
            INSERT INTO causacion_lineas (
                causacion_id,
                tipo,
                cuenta_contable,
                nombre_cuenta,
                centro_costo,
                debito,
                credito,
                concepto_servicio,
                descripcion
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            causacion_id,
            linea.get("tipo"),
            linea.get("cuenta_contable"),
            linea.get("nombre_cuenta"),
            linea.get("centro_costo"),
            linea.get("debito"),
            linea.get("credito"),
            linea.get("concepto_servicio"),
            linea.get("descripcion")
        ))

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "estado": "CAUSADA_SIMULADA",
        "mensaje": "Factura causada automÃ¡ticamente en modo simulaciÃ³n usando valores reales del XML.",
        "factura_id": factura_id,
        "causacion_id": causacion_id,
        "proveedor_nit": propuesta.get("proveedor_nit"),
        "proveedor_nombre": propuesta.get("proveedor_nombre"),
        "numero_factura": propuesta.get("numero_factura"),
        "confianza": propuesta.get("confianza"),
        "total_debito": total_debito,
        "total_credito": total_credito,
        "lineas": lineas
    }


@app.post("/revision/procesar-respuesta")
def procesar_respuesta_revision(payload: dict = Body(...)):
    subject = payload.get("subject", "")
    body = payload.get("body", "")
    from_email = payload.get("from", "")

    try:
        if "#CREAR_PROVEEDOR_SIIGO" in limpiar_cuerpo_correo(body):
            return procesar_respuesta_proveedor_siigo(payload)

        datos = parsear_respuesta_causacion(subject, body)

        factura_id = datos["factura_id"]
        lineas_usuario = datos["lineas"]

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT *
            FROM facturas_recibidas
            WHERE id = ?
        """, (factura_id,))

        factura_row = cursor.fetchone()

        if not factura_row:
            conn.close()
            return JSONResponse(
                status_code=404,
                content={
                    "ok": False,
                    "estado": "FACTURA_NO_ENCONTRADA",
                    "mensaje": f"No existe la factura_id {factura_id}."
                }
            )

        factura = dict(factura_row)

        cursor.execute("""
            SELECT *
            FROM facturas_lineas
            WHERE factura_id = ?
            ORDER BY id
        """, (factura_id,))

        lineas_factura = [dict(row) for row in cursor.fetchall()]

        cursor.execute("""
            SELECT id
            FROM causaciones
            WHERE factura_id = ?
              AND estado IN ('CAUSADA_SIMULADA', 'CAUSADA_RESPUESTA_MANUAL')
            LIMIT 1
        """, (factura_id,))

        causacion_existente = cursor.fetchone()

        if causacion_existente:
            conn.close()
            return {
                "ok": True,
                "estado": "YA_CAUSADA",
                "mensaje": "La factura ya tenÃ­a una causaciÃ³n registrada.",
                "factura_id": factura_id,
                "causacion_id": causacion_existente[0]
            }

        resultado_valores = construir_lineas_respuesta_con_valores_factura(
            factura,
            lineas_usuario
        )

        lineas = resultado_valores["lineas"]
        total_debito = resultado_valores["total_debito"]
        total_credito = resultado_valores["total_credito"]

        for linea in lineas:
            cursor.execute("""
                INSERT INTO contabilizaciones_historicas (
                    proveedor_nit,
                    proveedor_nombre,
                    descripcion,
                    cuenta_contable,
                    nombre_cuenta,
                    centro_costo,
                    debito,
                    credito,
                    cuenta_iva,
                    cuenta_retencion,
                    fuente,
                    fecha_documento
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                factura.get("proveedor_nit"),
                factura.get("proveedor_nombre"),
                linea.get("descripcion"),
                linea.get("cuenta_contable"),
                linea.get("nombre_cuenta"),
                linea.get("centro_costo"),
                linea.get("debito"),
                linea.get("credito"),
                linea.get("cuenta_contable") if linea.get("tipo") == "IVA" else None,
                None,
                f"Respuesta correo revisiÃ³n - {from_email}",
                factura.get("fecha_factura")
            ))

            upsert_mapeo_erp_desde_linea_manual(cursor, factura, linea)

        reglas_aprendidas = upsert_reglas_aprendidas_desde_respuesta_manual(
            cursor,
            factura,
            lineas_factura,
            lineas
        )

        cursor.execute("""
            INSERT INTO causaciones (
                factura_id,
                proveedor_nit,
                proveedor_nombre,
                numero_factura,
                estado,
                confianza,
                total_debito,
                total_credito,
                mensaje
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            factura_id,
            factura.get("proveedor_nit"),
            factura.get("proveedor_nombre"),
            factura.get("numero_factura"),
            "CAUSADA_RESPUESTA_MANUAL",
            100,
            total_debito,
            total_credito,
            "Factura causada desde respuesta manual por correo usando valores reales del XML."
        ))

        causacion_id = cursor.lastrowid

        for linea in lineas:
            cursor.execute("""
                INSERT INTO causacion_lineas (
                    causacion_id,
                    tipo,
                    cuenta_contable,
                    nombre_cuenta,
                    centro_costo,
                    debito,
                    credito,
                    concepto_servicio,
                    descripcion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                causacion_id,
                linea.get("tipo"),
                linea.get("cuenta_contable"),
                linea.get("nombre_cuenta"),
                linea.get("centro_costo"),
                linea.get("debito"),
                linea.get("credito"),
                linea.get("concepto_servicio"),
                linea.get("descripcion")
            ))

        conn.commit()
        conn.close()

        return {
            "ok": True,
            "estado": "CAUSADA_RESPUESTA_MANUAL",
            "mensaje": "La factura fue causada con la respuesta manual usando valores reales del XML y el histÃ³rico fue guardado.",
            "factura_id": factura_id,
            "causacion_id": causacion_id,
            "proveedor_nit": factura.get("proveedor_nit"),
            "proveedor_nombre": factura.get("proveedor_nombre"),
            "numero_factura": factura.get("numero_factura"),
            "total_debito": total_debito,
            "total_credito": total_credito,
            "total_pagar_factura": resultado_valores["total_pagar_factura"],
            "iva_factura": resultado_valores["iva_factura"],
            "valor_gasto_factura": resultado_valores["valor_gasto_factura"],
            "reglas_aprendidas": reglas_aprendidas,
            "total_reglas_aprendidas": len(reglas_aprendidas),
            "lineas": lineas
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "ERROR_RESPUESTA_REVISION",
                "mensaje": str(e)
            }
        )


@app.get("/causaciones/listar")
def listar_causaciones():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            factura_id,
            proveedor_nit,
            proveedor_nombre,
            numero_factura,
            estado,
            confianza,
            total_debito,
            total_credito,
            mensaje,
            siigo_comprobante_id,
            fecha_creacion
        FROM causaciones
        ORDER BY id DESC
        LIMIT 50
    """)

    causaciones = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return {
        "ok": True,
        "causaciones": causaciones
    }


@app.post("/facturas/procesar-adjunto")
async def procesar_adjunto(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    nombre_original = file.filename or "archivo_sin_nombre"
    extension = os.path.splitext(nombre_original)[1].lower()

    ruta_guardada = os.path.join(
        ADJUNTOS_PATH,
        f"{file_id}_{nombre_original}"
    )

    with open(ruta_guardada, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    carpeta_trabajo = os.path.join(PROCESADAS_PATH, file_id)
    os.makedirs(carpeta_trabajo, exist_ok=True)

    try:
        if extension == ".zip":
            with zipfile.ZipFile(ruta_guardada, "r") as zip_ref:
                safe_extract_zip(zip_ref, carpeta_trabajo)

            xml_files, pdf_files = buscar_xml_pdf(carpeta_trabajo)

        elif extension == ".xml":
            xml_files = [ruta_guardada]
            pdf_files = []

        else:
            return JSONResponse(
                status_code=400,
                content={
                    "ok": False,
                    "error": f"Tipo de archivo no soportado: {extension}",
                    "archivo_recibido": ruta_guardada
                }
            )

        if not xml_files:
            return JSONResponse(
                status_code=400,
                content={
                    "ok": False,
                    "error": "No se encontrÃ³ ningÃºn XML dentro del archivo recibido.",
                    "archivo_recibido": ruta_guardada,
                    "carpeta_trabajo": carpeta_trabajo
                }
            )

        xml_principal = xml_files[0]
        pdf_principal = pdf_files[0] if pdf_files else None

        datos_factura = extraer_datos_xml(xml_principal)
        validacion_nit = validar_nit_cliente_permitido(datos_factura)

        if not validacion_nit.get("permitido"):
            return {
                "ok": False,
                "estado": "CLIENTE_NIT_NO_PERMITIDO",
                "mensaje": (
                    "La factura no fue procesada porque el NIT del cliente/receptor "
                    "no esta en la lista NITS_CLIENTE_PERMITIDOS."
                ),
                "duplicada": False,
                "factura_id": None,
                "nombre_original": nombre_original,
                "archivo_recibido": ruta_guardada,
                "carpeta_trabajo": carpeta_trabajo,
                "total_xml_encontrados": len(xml_files),
                "total_pdf_encontrados": len(pdf_files),
                "xml_principal": xml_principal,
                "pdf_principal": pdf_principal,
                "factura": datos_factura,
                "cliente_nit": datos_factura.get("cliente_nit"),
                "cliente_nit_normalizado": validacion_nit.get("cliente_nit_normalizado"),
                "nits_permitidos": validacion_nit.get("nits_permitidos")
            }

        resultado_proveedor = registrar_proveedor(datos_factura)

        resultado_db = guardar_factura_db(
            datos_factura,
            xml_principal,
            pdf_principal,
            ruta_guardada
        )

        return {
            "ok": True,
            "mensaje": "Factura procesada correctamente",
            "factura_id": resultado_db.get("factura_id"),
            "duplicada": resultado_db.get("duplicada"),
            "proveedor_id": resultado_proveedor.get("proveedor_id"),
            "proveedor_nuevo": resultado_proveedor.get("proveedor_nuevo"),
            "nombre_original": nombre_original,
            "archivo_recibido": ruta_guardada,
            "carpeta_trabajo": carpeta_trabajo,
            "total_xml_encontrados": len(xml_files),
            "total_pdf_encontrados": len(pdf_files),
            "xml_principal": xml_principal,
            "pdf_principal": pdf_principal,
            "factura": datos_factura,
            "validacion_nit_cliente": validacion_nit
        }

    except zipfile.BadZipFile:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": "El archivo ZIP estÃ¡ daÃ±ado o no es un ZIP vÃ¡lido.",
                "archivo_recibido": ruta_guardada
            }
        )

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e),
                "archivo_recibido": ruta_guardada
            }
        )

@app.get("/siigo/auth-test")
def siigo_auth_test():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()
        resultado = client.autenticar()

        return {
            "ok": True,
            "enabled": client.enabled,
            "mensaje": "ConexiÃ³n con SIIGO validada correctamente.",
            "siigo": resultado
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "mensaje": "No fue posible autenticar contra SIIGO.",
                "error": str(e)
            }
        )


@app.get("/siigo/config")
def siigo_config():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        return {
            "ok": True,
            "enabled": client.enabled,
            "base_url": client.base_url,
            "username_configurado": bool(client.username),
            "access_key_configurado": bool(client.access_key),
            "partner_id": client.partner_id
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": str(e)
            }
        )

@app.get("/siigo/catalogos")
def siigo_catalogos():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        catalogos = client.consultar_catalogos_basicos()

        return {
            "ok": True,
            "mensaje": "CatÃ¡logos SIIGO consultados correctamente.",
            "catalogos": catalogos
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "mensaje": "No fue posible consultar catÃ¡logos de SIIGO.",
                "error": str(e)
            }
        )


@app.get("/siigo/documentos-compra")
def siigo_documentos_compra():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        return {
            "ok": True,
            "documentos": client.consultar_tipos_documento_compra()
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": str(e)
            }
        )


@app.get("/siigo/medios-pago")
def siigo_medios_pago():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        return {
            "ok": True,
            "medios_pago": client.consultar_medios_pago()
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": str(e)
            }
        )


@app.get("/siigo/impuestos")
def siigo_impuestos():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        return {
            "ok": True,
            "impuestos": client.consultar_impuestos()
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": str(e)
            }
        )


@app.get("/siigo/centros-costo")
def siigo_centros_costo():
    try:
        from siigo_client import SiigoClient

        client = SiigoClient()

        return {
            "ok": True,
            "centros_costo": client.consultar_centros_costo()
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": str(e)
            }
        )

def separar_prefijo_numero_siigo(numero_factura: str):
    numero_limpio = re.sub(r"[^A-Za-z0-9]", "", numero_factura or "")

    if not numero_limpio:
        return {
            "prefix": "FC",
            "number": "0"
        }

    match = re.search(r"(\d+)$", numero_limpio)

    if match:
        number = match.group(1)
        prefix = numero_limpio[:match.start()]
    else:
        prefix = "FC"
        number = re.sub(r"\D", "", numero_limpio) or "0"

    prefix = prefix[:6] if prefix else "FC"
    number = number[-11:] if len(number) > 11 else number

    return {
        "prefix": prefix,
        "number": number
    }


def calcular_fecha_vencimiento_siigo(fecha_factura: str):
    try:
        from datetime import datetime, timedelta

        dias = int(os.getenv("SIIGO_DIAS_VENCIMIENTO_COMPRA", "30"))
        fecha = datetime.strptime(fecha_factura, "%Y-%m-%d")
        return (fecha + timedelta(days=dias)).strftime("%Y-%m-%d")

    except Exception:
        return fecha_factura


def obtener_env_int(nombre: str):
    valor = os.getenv(nombre)

    if valor is None or str(valor).strip() == "":
        raise ValueError(f"Falta configurar {nombre} en el archivo .env")

    return int(valor)


def obtener_env_texto(nombre, default=None):
    valor = os.getenv(nombre)

    if valor is None:
        return default

    valor = str(valor).strip()
    return valor if valor else default


def calcular_digito_verificacion_nit(nit: str | None):
    nit_normalizado = normalizar_nit(nit)

    if not nit_normalizado:
        return None

    pesos = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    digitos = [int(d) for d in nit_normalizado if d.isdigit()]
    pesos_usar = pesos[-len(digitos):]
    suma = sum(d * p for d, p in zip(digitos, pesos_usar))
    residuo = suma % 11

    if residuo in [0, 1]:
        return str(residuo)

    return str(11 - residuo)


def extraer_datos_proveedor_siigo_desde_xml(xml_path: str | None):
    datos = {}

    if not xml_path or not os.path.exists(xml_path):
        return datos

    try:
        root = cargar_xml_factura(xml_path)
    except Exception:
        return datos

    ns = {
        "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
    }

    party_path = ".//cac:AccountingSupplierParty//cac:Party"

    datos["nombre"] = (
        texto(root, party_path + "//cac:PartyLegalEntity//cbc:RegistrationName", ns)
        or texto(root, party_path + "//cac:PartyName//cbc:Name", ns)
    )
    datos["nit"] = texto(root, party_path + "//cac:PartyTaxScheme//cbc:CompanyID", ns)
    datos["check_digit"] = texto_atributo(
        root,
        party_path + "//cac:PartyTaxScheme//cbc:CompanyID",
        "schemeID",
        ns
    )

    address_paths = [
        party_path + "//cac:PhysicalLocation//cac:Address",
        party_path + "//cac:PartyLegalEntity//cac:RegistrationAddress",
        party_path + "//cac:PartyTaxScheme//cac:RegistrationAddress"
    ]

    for address_path in address_paths:
        direccion = (
            texto(root, address_path + "//cac:AddressLine//cbc:Line", ns)
            or texto(root, address_path + "/cbc:StreetName", ns)
        )
        ciudad_codigo = texto(root, address_path + "/cbc:ID", ns)
        ciudad_nombre = texto(root, address_path + "/cbc:CityName", ns)
        departamento_codigo = texto(root, address_path + "/cbc:CountrySubentityCode", ns)
        departamento_nombre = texto(root, address_path + "/cbc:CountrySubentity", ns)
        pais_codigo = texto(root, address_path + "//cac:Country/cbc:IdentificationCode", ns)

        if direccion and not datos.get("direccion"):
            datos["direccion"] = direccion

        if ciudad_codigo and not datos.get("ciudad_codigo"):
            datos["ciudad_codigo"] = ciudad_codigo

        if ciudad_nombre and not datos.get("ciudad_nombre"):
            datos["ciudad_nombre"] = ciudad_nombre

        if departamento_codigo and not datos.get("departamento_codigo"):
            datos["departamento_codigo"] = departamento_codigo

        if departamento_nombre and not datos.get("departamento_nombre"):
            datos["departamento_nombre"] = departamento_nombre

        if pais_codigo and not datos.get("pais_codigo"):
            datos["pais_codigo"] = pais_codigo.upper()

    datos["email"] = texto(root, party_path + "//cac:Contact//cbc:ElectronicMail", ns)
    datos["telefono"] = texto(root, party_path + "//cac:Contact//cbc:Telephone", ns)

    return {k: v for k, v in datos.items() if v not in [None, ""]}


def consultar_proveedor_siigo(client, nit: str):
    nit_normalizado = normalizar_nit(nit)

    if not nit_normalizado:
        return None

    respuesta = client.get(
        "/v1/customers",
        params={
            "identification": nit_normalizado,
            "branch_office": 0,
            "page": 1,
            "page_size": 25
        }
    )

    if isinstance(respuesta, list):
        candidatos = respuesta
    else:
        candidatos = (
            respuesta.get("results")
            or respuesta.get("data")
            or respuesta.get("items")
            or []
        )

    for candidato in candidatos:
        identificacion = normalizar_nit(candidato.get("identification"))
        sucursal = candidato.get("branch_office", 0)

        if identificacion == nit_normalizado and int(sucursal or 0) == 0:
            return candidato

    return None


def construir_datos_proveedor_siigo(factura: dict, extras: dict | None = None):
    extras = extras or {}
    xml_datos = extraer_datos_proveedor_siigo_desde_xml(factura.get("xml_principal"))

    def tomar(clave, *envs):
        for origen in [extras, xml_datos, factura]:
            valor = origen.get(clave)
            if valor not in [None, ""]:
                return str(valor).strip()

        for env_name in envs:
            valor = obtener_env_texto(env_name)
            if valor:
                return valor

        return None

    nit = normalizar_nit(tomar("nit") or factura.get("proveedor_nit"))
    nombre = tomar("nombre") or factura.get("proveedor_nombre")
    direccion = tomar("direccion", "SIIGO_PROVEEDOR_DIRECCION_DEFAULT")
    pais_codigo = tomar("pais_codigo", "SIIGO_PROVEEDOR_PAIS_DEFAULT") or "CO"
    departamento_codigo = tomar("departamento_codigo", "SIIGO_PROVEEDOR_DEPARTAMENTO_DEFAULT")
    ciudad_codigo = tomar("ciudad_codigo", "SIIGO_PROVEEDOR_CIUDAD_DEFAULT")
    responsabilidad = tomar("responsabilidad_fiscal", "SIIGO_PROVEEDOR_RESPONSABILIDAD_DEFAULT") or "R-99-PN"
    telefono = tomar("telefono", "SIIGO_PROVEEDOR_TELEFONO_DEFAULT")
    email = tomar("email", "SIIGO_PROVEEDOR_EMAIL_DEFAULT")
    check_digit = tomar("check_digit") or calcular_digito_verificacion_nit(nit)

    faltantes = []

    if not nit:
        faltantes.append("NIT")

    if not nombre:
        faltantes.append("NOMBRE")

    if not direccion:
        faltantes.append("DIRECCION")

    if not departamento_codigo:
        faltantes.append("DEPARTAMENTO_CODIGO")

    if not ciudad_codigo:
        faltantes.append("CIUDAD_CODIGO")

    return {
        "nit": nit,
        "nombre": nombre,
        "check_digit": check_digit,
        "direccion": direccion,
        "pais_codigo": pais_codigo,
        "departamento_codigo": departamento_codigo,
        "ciudad_codigo": ciudad_codigo,
        "responsabilidad_fiscal": responsabilidad,
        "telefono": telefono,
        "email": email,
        "faltantes": faltantes,
        "xml_datos": xml_datos,
        "extras": extras
    }


def construir_payload_proveedor_siigo(datos: dict):
    payload = {
        "type": "Supplier",
        "person_type": "Company",
        "id_type": "31",
        "identification": datos.get("nit"),
        "name": [
            datos.get("nombre")
        ],
        "commercial_name": datos.get("nombre"),
        "branch_office": 0,
        "active": True,
        "vat_responsible": True,
        "fiscal_responsibilities": [
            {
                "code": datos.get("responsabilidad_fiscal") or "R-99-PN"
            }
        ],
        "address": {
            "address": datos.get("direccion"),
            "city": {
                "country_code": datos.get("pais_codigo") or "CO",
                "state_code": datos.get("departamento_codigo"),
                "city_code": datos.get("ciudad_codigo")
            }
        },
        "comments": "Creado automaticamente desde FacturasIA."
    }

    if datos.get("check_digit"):
        payload["check_digit"] = str(datos.get("check_digit"))

    telefono = re.sub(r"\D+", "", str(datos.get("telefono") or ""))

    if telefono:
        payload["phones"] = [
            {
                "indicative": "57",
                "number": telefono[:10]
            }
        ]

    if datos.get("email"):
        payload["contacts"] = [
            {
                "first_name": datos.get("nombre")[:50],
                "last_name": "Proveedor",
                "email": datos.get("email")
            }
        ]

    return payload


def respuesta_requiere_datos_proveedor(causacion_id: int, factura: dict, datos: dict):
    mensaje = "El proveedor no existe en SIIGO y faltan datos para crearlo automaticamente."

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE causaciones
        SET estado = ?, mensaje = ?
        WHERE id = ?
    """, (
        "REQUIERE_DATOS_PROVEEDOR_SIIGO",
        mensaje,
        causacion_id
    ))
    conn.commit()
    conn.close()

    return JSONResponse(
        status_code=400,
        content={
            "ok": False,
            "estado": "REQUIERE_DATOS_PROVEEDOR_SIIGO",
            "mensaje": mensaje,
            "causacion_id": causacion_id,
            "factura_id": factura.get("id"),
            "proveedor_nombre": factura.get("proveedor_nombre"),
            "proveedor_nit": factura.get("proveedor_nit"),
            "numero_factura": factura.get("numero_factura"),
            "datos_proveedor": {
                k: v for k, v in datos.items()
                if k not in {"xml_datos", "extras"}
            },
            "datos_faltantes": datos.get("faltantes") or [],
            "formato_respuesta": (
                "#CREAR_PROVEEDOR_SIIGO\n\n"
                f"CAUSACION_ID={causacion_id}\n"
                f"NOMBRE={factura.get('proveedor_nombre') or ''}\n"
                f"NIT={factura.get('proveedor_nit') or ''}\n"
                "DIRECCION=\n"
                "DEPARTAMENTO_CODIGO=\n"
                "CIUDAD_CODIGO=\n"
                "PAIS_CODIGO=CO\n"
                "RESPONSABILIDAD_FISCAL=R-99-PN\n"
                "TELEFONO=\n"
                "EMAIL=\n"
            )
        }
    )


def asegurar_proveedor_siigo(client, factura: dict, causacion_id: int, extras: dict | None = None):
    if not SIIGO_AUTO_CREAR_PROVEEDOR:
        return {
            "ok": True,
            "auto_creacion_habilitada": False,
            "mensaje": "Creacion automatica de proveedor deshabilitada."
        }

    existente = consultar_proveedor_siigo(client, factura.get("proveedor_nit"))

    if existente:
        return {
            "ok": True,
            "proveedor_existia": True,
            "proveedor_siigo": existente
        }

    datos = construir_datos_proveedor_siigo(factura, extras)

    if datos.get("faltantes"):
        return {
            "ok": False,
            "requiere_datos": True,
            "datos": datos
        }

    payload = construir_payload_proveedor_siigo(datos)
    respuesta = client.post("/v1/customers", payload)

    return {
        "ok": True,
        "proveedor_creado": True,
        "proveedor_siigo": respuesta,
        "payload_proveedor": payload
    }


def ajustar_redondeo_items_siigo(items: list, total_pagar: float, iva: float):
    if not items or iva <= 0 or total_pagar <= 0:
        return None

    items_gravados = [
        item
        for item in items
        if item.get("taxes")
    ]

    if not items_gravados:
        return None

    base_gravada = round(
        sum(float(item.get("price") or 0) for item in items_gravados),
        2
    )

    if base_gravada <= 0:
        return None

    try:
        tasa = float(os.getenv("SIIGO_TASA_IVA_19", "0.19"))
    except ValueError:
        tasa = 0.19

    if tasa > 1:
        tasa = tasa / 100

    if tasa <= 0:
        tasa = 0.19

    def calcular_total():
        total = 0

        for item in items:
            precio = float(item.get("price") or 0)

            if item.get("taxes"):
                total += precio * (1 + tasa)
            else:
                total += precio

        return round(total, 2)

    total_calculado = calcular_total()
    diferencia = round(total_calculado - total_pagar, 2)

    if not diferencia:
        return {
            "aplicado": False,
            "total_calculado": total_calculado,
            "total_pagar": total_pagar,
            "diferencia": diferencia,
            "motivo": "No requiere ajuste de redondeo."
        }

    try:
        max_ajuste = float(os.getenv("SIIGO_MAX_AJUSTE_REDONDEO", "1.00"))
    except ValueError:
        max_ajuste = 1.00

    if max_ajuste < 0:
        max_ajuste = 0

    if abs(diferencia) > max_ajuste:
        return {
            "aplicado": False,
            "total_calculado": total_calculado,
            "total_pagar": total_pagar,
            "diferencia": diferencia,
            "max_ajuste": max_ajuste,
            "motivo": "Diferencia fuera del rango de redondeo automatico."
        }

    item_ajustado = items_gravados[-1]
    precio_original = round(float(item_ajustado.get("price") or 0), 2)
    ajuste_base = diferencia / (1 + tasa)
    precio_ajustado = round(precio_original - ajuste_base, 2)

    if precio_ajustado <= 0:
        return {
            "aplicado": False,
            "total_calculado": total_calculado,
            "total_pagar": total_pagar,
            "diferencia": diferencia,
            "motivo": "El ajuste dejaria el precio del item en cero o negativo."
        }

    item_ajustado["price"] = precio_ajustado
    total_ajustado = calcular_total()

    return {
        "aplicado": True,
        "tasa": round(tasa, 6),
        "precio_original": precio_original,
        "precio_ajustado": precio_ajustado,
        "total_calculado": total_calculado,
        "total_ajustado": total_ajustado,
        "total_pagar": total_pagar,
        "diferencia": diferencia,
        "max_ajuste": max_ajuste
    }


def construir_payload_siigo_compra_desde_causacion(causacion_id: int):
    document_id = obtener_env_int("SIIGO_DOCUMENT_ID_COMPRA")
    payment_id = obtener_env_int("SIIGO_PAYMENT_ID_COMPRA")
    tax_id_iva_19 = obtener_env_int("SIIGO_TAX_ID_IVA_19")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM causaciones
        WHERE id = ?
    """, (causacion_id,))

    causacion_row = cursor.fetchone()

    if not causacion_row:
        conn.close()
        raise ValueError(f"No existe la causaciÃ³n {causacion_id}.")

    causacion = dict(causacion_row)

    cursor.execute("""
        SELECT *
        FROM facturas_recibidas
        WHERE id = ?
    """, (causacion.get("factura_id"),))

    factura_row = cursor.fetchone()

    if not factura_row:
        conn.close()
        raise ValueError(f"No existe la factura asociada a la causaciÃ³n {causacion_id}.")

    factura = dict(factura_row)

    cursor.execute("""
        SELECT *
        FROM causacion_lineas
        WHERE causacion_id = ?
        ORDER BY id
    """, (causacion_id,))

    lineas = [dict(row) for row in cursor.fetchall()]

    total_pagar = round(float(factura.get("total_pagar") or 0), 2)
    iva = round(float(factura.get("iva") or 0), 2)

    lineas_gasto = [
        l for l in lineas
        if float(l.get("debito") or 0) > 0
        and str(l.get("tipo") or "").upper() not in ["IVA", "CXP", "CUENTA_POR_PAGAR"]
    ]

    if not lineas_gasto:
        conn.close()
        raise ValueError("No se encontro una linea de gasto para construir el item SIIGO.")

    items = []
    mapeos_usados = []

    for linea_gasto in lineas_gasto:
        concepto = linea_gasto.get("concepto_servicio")

        if not concepto:
            conn.close()
            raise ValueError(
                "REQUIERE_REVISION_CONCEPTO: una linea de gasto no tiene concepto_servicio."
            )

        mapeo = buscar_mapeo_erp(
            cursor,
            factura.get("proveedor_nit"),
            concepto,
            linea_gasto.get("cuenta_contable")
        )

        if not mapeo:
            conn.close()
            raise ValueError(
                f"REQUIERE_MAPEO_ERP: no existe mapeo ERP para proveedor "
                f"{factura.get('proveedor_nit')} y concepto {concepto}."
            )

        item_type = mapeo.get("item_type_erp")
        item_code = mapeo.get("item_code_erp")

        if not item_type or not item_code:
            conn.close()
            raise ValueError(
                f"REQUIERE_MAPEO_ERP: el mapeo ERP {mapeo.get('id')} no tiene "
                "item_type_erp o item_code_erp."
            )

        item = {
            "type": item_type,
            "code": str(item_code),
            "description": (
                mapeo.get("item_description_erp")
                or linea_gasto.get("descripcion")
                or linea_gasto.get("nombre_cuenta")
                or "Gasto factura proveedor"
            ),
            "quantity": 1,
            "price": round(float(linea_gasto.get("debito") or 0), 2)
        }

        tax_id = mapeo.get("tax_id_erp") or tax_id_iva_19

        if iva > 0 and tax_id:
            item["taxes"] = [
                {
                    "id": int(tax_id)
                }
            ]

        items.append(item)
        mapeos_usados.append({
            "mapeo_erp_id": mapeo.get("id"),
            "concepto_servicio": concepto,
            "item_type_erp": item_type,
            "item_code_erp": item_code,
            "document_id_erp": mapeo.get("document_id_erp"),
            "payment_id_erp": mapeo.get("payment_id_erp"),
            "tax_id_erp": mapeo.get("tax_id_erp")
        })

    primer_mapeo = mapeos_usados[0] if mapeos_usados else {}
    document_id = int(primer_mapeo.get("document_id_erp") or document_id)
    payment_id = int(primer_mapeo.get("payment_id_erp") or payment_id)
    ajuste_redondeo_siigo = ajustar_redondeo_items_siigo(items, total_pagar, iva)
    conn.close()

    datos_factura_proveedor = separar_prefijo_numero_siigo(
        factura.get("numero_factura")
    )

    payload = {
        "document": {
            "id": document_id
        },
        "date": factura.get("fecha_factura"),
        "supplier": {
            "identification": str(factura.get("proveedor_nit")),
            "branch_office": 0
        },
        "provider_invoice": {
            "prefix": datos_factura_proveedor["prefix"],
            "number": datos_factura_proveedor["number"]
        },
        "tax_included": False,
        "observations": (
            f"Factura generada desde FacturasIA. "
            f"Factura ID: {factura.get('id')}. "
            f"Causacion ID: {causacion_id}. "
            f"Factura proveedor: {factura.get('numero_factura')}."
        ),
        "items": items,
        "payments": [
            {
                "id": payment_id,
                "value": total_pagar,
                "due_date": calcular_fecha_vencimiento_siigo(factura.get("fecha_factura"))
            }
        ]
    }

    return {
        "payload": payload,
        "factura": factura,
        "causacion": causacion,
        "lineas": lineas,
        "validacion": {
            "total_pagar_factura": total_pagar,
            "iva_factura": iva,
            "document_id": document_id,
            "payment_id": payment_id,
            "tax_id_iva_19": tax_id_iva_19,
            "mapeos_usados": mapeos_usados,
            "items_generados": len(items),
            "ajuste_redondeo_siigo": ajuste_redondeo_siigo,
            "modo": "SIMULACION_NO_ENVIADO"
        }
    }

    conn.close()

    total_pagar = round(float(factura.get("total_pagar") or 0), 2)
    iva = round(float(factura.get("iva") or 0), 2)
    subtotal = round(float(factura.get("subtotal") or 0), 2)
    total_sin_impuestos = round(float(factura.get("total_sin_impuestos") or 0), 2)

    if subtotal > 0:
       valor_gasto_factura = subtotal
    elif total_sin_impuestos > 0:
       valor_gasto_factura = total_sin_impuestos
    else:
       valor_gasto_factura = round(total_pagar - iva, 2)

    lineas_gasto = [
        l for l in lineas
        if float(l.get("debito") or 0) > 0
        and str(l.get("tipo") or "").upper() not in ["IVA"]
        and str(l.get("cuenta_contable") or "") != "240805"
    ]

    if not lineas_gasto:
        raise ValueError("No se encontrÃ³ una lÃ­nea de gasto para construir el Ã­tem SIIGO.")

    if len(lineas_gasto) > 1:
        raise ValueError(
            "La simulaciÃ³n SIIGO actual solo soporta una lÃ­nea de gasto. "
            "Luego habilitamos mÃºltiples lÃ­neas."
        )

    linea_gasto = lineas_gasto[0]

    datos_factura_proveedor = separar_prefijo_numero_siigo(
        factura.get("numero_factura")
    )

    item = {
        "type": "Account",
        "code": str(linea_gasto.get("cuenta_contable")),
        "description": linea_gasto.get("nombre_cuenta") or linea_gasto.get("descripcion") or "Gasto factura proveedor",
        "quantity": 1,
        "price": valor_gasto_factura
    }

    if iva > 0:
        item["taxes"] = [
            {
                "id": tax_id_iva_19
            }
        ]

    payload = {
        "document": {
            "id": document_id
        },
        "date": factura.get("fecha_factura"),
        "supplier": {
            "identification": str(factura.get("proveedor_nit")),
            "branch_office": 0
        },
        "provider_invoice": {
            "prefix": datos_factura_proveedor["prefix"],
            "number": datos_factura_proveedor["number"]
        },
        "tax_included": False,
        "observations": (
            f"Factura generada desde FacturasIA. "
            f"Factura ID: {factura.get('id')}. "
            f"CausaciÃ³n ID: {causacion_id}. "
            f"Factura proveedor: {factura.get('numero_factura')}."
        ),
        "items": [
            item
        ],
        "payments": [
            {
                "id": payment_id,
                "value": total_pagar,
                "due_date": calcular_fecha_vencimiento_siigo(factura.get("fecha_factura"))
            }
        ]
    }

    return {
        "payload": payload,
        "factura": factura,
        "causacion": causacion,
        "lineas": lineas,
        "validacion": {
            "total_pagar_factura": total_pagar,
            "iva_factura": iva,
            "cuenta_gasto_siigo": linea_gasto.get("cuenta_contable"),
            "valor_gasto_factura": valor_gasto_factura,
            "document_id": document_id,
            "payment_id": payment_id,
            "tax_id_iva_19": tax_id_iva_19,
            "modo": "SIMULACION_NO_ENVIADO"
        }
    }


@app.get("/siigo/preparar-compra/{causacion_id}")
def siigo_preparar_compra(causacion_id: int):
    try:
        resultado = construir_payload_siigo_compra_desde_causacion(causacion_id)

        return {
            "ok": True,
            "mensaje": "Payload de compra SIIGO preparado correctamente. No se enviÃ³ a SIIGO.",
            "modo": "SIMULACION_NO_ENVIADO",
            "causacion_id": causacion_id,
            "validacion": resultado["validacion"],
            "payload_siigo": resultado["payload"]
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "mensaje": "No fue posible preparar el payload SIIGO.",
                "error": str(e),
                "causacion_id": causacion_id
            }
        )

@app.post("/siigo/enviar-compra/{causacion_id}")
def siigo_enviar_compra(causacion_id: int):
    try:
        from siigo_client import SiigoClient

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                id,
                estado,
                siigo_comprobante_id
            FROM causaciones
            WHERE id = ?
        """, (causacion_id,))

        causacion_row = cursor.fetchone()
        conn.close()

        if not causacion_row:
            return JSONResponse(
                status_code=404,
                content={
                    "ok": False,
                    "estado": "CAUSACION_NO_ENCONTRADA",
                    "mensaje": f"No existe la causacion {causacion_id}.",
                    "causacion_id": causacion_id
                }
            )

        causacion_actual = dict(causacion_row)

        if (
            causacion_actual.get("estado") == "ENVIADA_SIIGO"
            or causacion_actual.get("siigo_comprobante_id")
        ):
            return {
                "ok": True,
                "estado": "YA_ENVIADA_SIIGO",
                "mensaje": "La causacion ya fue enviada a SIIGO. No se realizo un nuevo envio.",
                "causacion_id": causacion_id,
                "siigo_id": causacion_actual.get("siigo_comprobante_id")
            }

        if os.getenv("SIIGO_ENABLED", "false").lower() != "true":
            return JSONResponse(
                status_code=400,
                content={
                    "ok": False,
                    "estado": "SIIGO_DESHABILITADO",
                    "mensaje": "SIIGO_ENABLED estÃ¡ en false. Cambie SIIGO_ENABLED=true en .env para permitir envÃ­o real.",
                    "causacion_id": causacion_id
                }
            )

        resultado = construir_payload_siigo_compra_desde_causacion(causacion_id)
        payload = resultado["payload"]

        client = SiigoClient()
        proveedor = asegurar_proveedor_siigo(
            client,
            resultado["factura"],
            causacion_id
        )

        if proveedor.get("requiere_datos"):
            return respuesta_requiere_datos_proveedor(
                causacion_id,
                resultado["factura"],
                proveedor.get("datos") or {}
            )

        respuesta_siigo = client.post("/v1/purchases", payload)

        siigo_id = (
            respuesta_siigo.get("id")
            or respuesta_siigo.get("uuid")
            or respuesta_siigo.get("purchase_id")
        )

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE causaciones
            SET
                estado = ?,
                siigo_comprobante_id = ?,
                mensaje = ?
            WHERE id = ?
        """, (
            "ENVIADA_SIIGO",
            str(siigo_id) if siigo_id else None,
            "Compra enviada correctamente a SIIGO.",
            causacion_id
        ))

        conn.commit()
        conn.close()

        return {
            "ok": True,
            "estado": "ENVIADA_SIIGO",
            "mensaje": "Compra enviada correctamente a SIIGO.",
            "causacion_id": causacion_id,
            "siigo_id": siigo_id,
            "respuesta_siigo": respuesta_siigo,
            "payload_enviado": payload
        }

    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "estado": "ERROR_ENVIANDO_SIIGO",
                "mensaje": str(e),
                "causacion_id": causacion_id
            }
        )

inicializar_db()
